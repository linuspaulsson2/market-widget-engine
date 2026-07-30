"""
Market Widget Data Updater
==========================
Hämtar aktiedata, beräknar nyckeltal och publicerar JSON till GitHub Gist.
Körs automatiskt via GitHub Actions var 30:e minut under börsdagar.
"""

import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import anthropic
import openpyxl
import pytz
import requests
import yfinance as yf

# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).parent
CONFIG_PATH = SCRIPT_DIR / "config.json"

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    CONFIG = json.load(f)

HOLDINGS = CONFIG["holdings"]
FOCUS_STOCKS = CONFIG.get("focus_stocks", [])

# Override focus stocks from focus_stocks.json if it exists (written by the iOS/macOS app).
# Falls back to config.json["focus_stocks"] if the file is missing.
_FOCUS_STOCKS_PATH = SCRIPT_DIR / "focus_stocks.json"
if _FOCUS_STOCKS_PATH.exists():
    try:
        with open(_FOCUS_STOCKS_PATH, "r", encoding="utf-8") as _f:
            _fs_data = json.load(_f)
        _focus_tickers = [t.strip().upper() for t in _fs_data.get("tickers", []) if t.strip()]
        if _focus_tickers:
            _known = {h["ticker"].upper(): h for h in CONFIG.get("holdings", []) + CONFIG.get("focus_stocks", [])}
            FOCUS_STOCKS = []
            for _t in _focus_tickers:
                if _t in _known:
                    FOCUS_STOCKS.append(_known[_t])
                else:
                    _suffix = _t.split(".")[-1] if "." in _t else ""
                    _cur = "SEK" if _suffix in ("ST", "HE", "CO") else "USD"
                    FOCUS_STOCKS.append({"ticker": _t, "name": _t, "currency": _cur})
    except Exception as _e:
        print(f"  VARNING: Kunde inte läsa focus_stocks.json: {_e}")

ALERT_THRESHOLD = CONFIG["alert_threshold_pct"]

# Tickers att hämta (aktier + USD/SEK)
STOCK_TICKERS = [h["ticker"] for h in HOLDINGS]
FX_TICKER = "USDSEK=X"

# Tidszoner
TZ_STOCKHOLM = pytz.timezone("Europe/Stockholm")
TZ_US_EAST = pytz.timezone("US/Eastern")

# Svenska veckodagar
WEEKDAY_SV = ["mån", "tis", "ons", "tor", "fre", "lör", "sön"]


# ---------------------------------------------------------------------------
# Hjälpfunktioner
# ---------------------------------------------------------------------------

def _us_market_holidays(year: int) -> set:
    """Returnera datum för fasta + rörliga US-börshelgdagar."""
    from datetime import date as _date
    holidays = set()
    # Nyårsdagen
    holidays.add(_date(year, 1, 1))
    # MLK Day – tredje måndagen i januari
    d = _date(year, 1, 1)
    mon_count = 0
    while mon_count < 3:
        d += timedelta(days=1)
        if d.weekday() == 0:
            mon_count += 1
    holidays.add(d)
    # Presidents Day – tredje måndagen i februari
    d = _date(year, 2, 1)
    mon_count = 0
    while mon_count < 3:
        d += timedelta(days=1)
        if d.weekday() == 0:
            mon_count += 1
    holidays.add(d)
    # Good Friday – påskfredag (anonymus Gauss-algoritm)
    a = year % 19
    b = year // 100
    c = year % 100
    d_val = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d_val - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    easter = _date(year, month, day)
    holidays.add(easter - timedelta(days=2))  # Good Friday
    # Memorial Day – sista måndagen i maj
    d = _date(year, 5, 31)
    while d.weekday() != 0:
        d -= timedelta(days=1)
    holidays.add(d)
    # Juneteenth
    holidays.add(_date(year, 6, 19))
    # Independence Day
    holidays.add(_date(year, 7, 4))
    # Labor Day – första måndagen i september
    d = _date(year, 9, 1)
    while d.weekday() != 0:
        d += timedelta(days=1)
    holidays.add(d)
    # Thanksgiving – fjärde torsdagen i november
    d = _date(year, 11, 1)
    thu_count = 0
    while thu_count < 4:
        if d.weekday() == 3:
            thu_count += 1
            if thu_count == 4:
                break
        d += timedelta(days=1)
    holidays.add(d)
    # Julafton
    holidays.add(_date(year, 12, 25))
    return holidays


def _se_market_holidays(year: int) -> set:
    """Returnera datum för svenska börshelgdagar."""
    from datetime import date as _date
    holidays = set()
    holidays.add(_date(year, 1, 1))   # Nyårsdagen
    holidays.add(_date(year, 1, 6))   # Trettondedag jul
    # Påsk (Gauss)
    a = year % 19
    b = year // 100
    c = year % 100
    d_val = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d_val - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    easter = _date(year, month, day)
    holidays.add(easter - timedelta(days=2))  # Långfredagen
    holidays.add(easter + timedelta(days=1))  # Annandag påsk
    holidays.add(_date(year, 5, 1))   # Första maj
    holidays.add(easter + timedelta(days=39))  # Kristi himmelsfärd
    holidays.add(_date(year, 6, 6))   # Nationaldagen
    # Midsommarafton – fredagen 19-25 juni
    d = _date(year, 6, 19)
    while d.weekday() != 4:
        d += timedelta(days=1)
    holidays.add(d)
    holidays.add(_date(year, 12, 24))  # Julafton
    holidays.add(_date(year, 12, 25))  # Juldagen
    holidays.add(_date(year, 12, 26))  # Annandag jul
    holidays.add(_date(year, 12, 31))  # Nyårsafton
    return holidays


def is_market_open() -> dict:
    """Kolla om Stockholmsbörsen och/eller USA-börsen är öppna just nu."""
    now_sthlm = datetime.now(TZ_STOCKHOLM)
    now_us = datetime.now(TZ_US_EAST)
    weekday_se = now_sthlm.weekday()  # 0=mån, 6=sön
    weekday_us = now_us.weekday()

    se_holiday = now_sthlm.date() in _se_market_holidays(now_sthlm.year)
    us_holiday = now_us.date() in _us_market_holidays(now_us.year)

    stockholm_open = (
        not se_holiday
        and weekday_se < 5
        and (9 <= now_sthlm.hour < 17
             or (now_sthlm.hour == 17 and now_sthlm.minute <= 30))
    )
    us_open = (
        not us_holiday
        and weekday_us < 5
        and (now_us.hour > 9 or (now_us.hour == 9 and now_us.minute >= 30))
        and now_us.hour < 16
    )

    return {
        "stockholm": stockholm_open,
        "us": us_open,
        "any_open": stockholm_open or us_open,
        "timestamp_swe": now_sthlm.strftime("%Y-%m-%d %H:%M"),
    }


def fetch_stock_data(holdings_list=None) -> list:
    """
    Hämta aktuell kursdata, ATH och 52-veckors-high för alla aktier.
    Använder yf.download() för att hämta allt på en gång (snabbare).
    """
    if holdings_list is None:
        holdings_list = HOLDINGS
    stocks = []

    for holding in holdings_list:
        ticker_str = holding["ticker"]
        name = holding["name"]
        currency = holding["currency"]

        try:
            ticker = yf.Ticker(ticker_str)

            # Hämta senaste 5 dagars data för aktuellt pris + förändring
            hist_5d = ticker.history(period="5d")
            if hist_5d.empty:
                print(f"  VARNING: Ingen data för {ticker_str}, hoppar över")
                continue

            current_price = hist_5d["Close"].iloc[-1]

            # Föregående stängning (näst sista dagen med data)
            if len(hist_5d) >= 2:
                previous_close = hist_5d["Close"].iloc[-2]
            else:
                previous_close = current_price

            # Daglig förändring
            change = current_price - previous_close
            change_pct = (change / previous_close * 100) if previous_close else 0

            # 52-veckors high
            hist_1y = ticker.history(period="1y")
            high_52w = hist_1y["High"].max() if not hist_1y.empty else current_price
            high_52w_pct = ((current_price - high_52w) / high_52w * 100) if high_52w else 0

            # All-time high (hämta max historik)
            hist_max = ticker.history(period="max")
            ath = hist_max["High"].max() if not hist_max.empty else current_price
            ath_pct = ((current_price - ath) / ath * 100) if ath else 0

            # MA200 + RSI(14) — beräknas från max-historik
            above_ma200 = None
            rsi = None
            try:
                closes = hist_max["Close"].dropna()
                if len(closes) >= 200:
                    ma200_val = closes.rolling(200).mean().iloc[-1]
                    above_ma200 = bool(float(current_price) > float(ma200_val))
                if len(closes) >= 15:
                    delta = closes.diff()
                    gain = delta.clip(lower=0).rolling(14).mean().iloc[-1]
                    loss = (-delta.clip(upper=0)).rolling(14).mean().iloc[-1]
                    if loss > 0:
                        rsi = round(100 - 100 / (1 + gain / loss), 1)
                    elif gain > 0:
                        rsi = 100.0
                    else:
                        rsi = 50.0
            except Exception:
                pass

            # Nyheter för denna aktie
            news = []
            try:
                ticker_news = ticker.news or []
                for item in ticker_news[:8]:
                    if isinstance(item, dict):
                        # Navigera yfinance:s nyhetsformat (kan vara nested)
                        content = item.get("content", {})
                        if isinstance(content, dict):
                            title = str(content.get("title", ""))
                            provider = content.get("provider", {})
                            provider_name = str(provider.get("displayName", "")) if isinstance(provider, dict) else ""
                            canon = content.get("canonicalUrl", {})
                            link = str(canon.get("url", "")) if isinstance(canon, dict) else ""
                        else:
                            title = str(item.get("title", "Ingen rubrik"))
                            provider_name = str(item.get("publisher", ""))
                            link = str(item.get("link", ""))

                        if title:
                            news.append({
                                "title": title,
                                "publisher": provider_name,
                                "link": link,
                            })
            except Exception as e:
                print(f"  Kunde inte hämta nyheter för {ticker_str}: {e}")

            # Kolla om det är en stor rörelse (alert)
            is_alert = abs(change_pct) >= ALERT_THRESHOLD

            stocks.append({
                "ticker": ticker_str,
                "name": name,
                "price": round(float(current_price), 2),
                "change": round(float(change), 2),
                "change_pct": round(float(change_pct), 2),
                "high_52w": round(float(high_52w), 2),
                "high_52w_pct": round(float(high_52w_pct), 1),
                "ath": round(float(ath), 2),
                "ath_pct": round(float(ath_pct), 1),
                "currency": currency,
                "is_alert": is_alert,
                "news": news,
                "domain": resolve_domain(ticker_str),
                "above_ma200": above_ma200,
                "rsi": rsi,
            })
            print(f"  OK: {name} ({ticker_str}) = {current_price:.2f} {currency} ({change_pct:+.2f}%) RSI={rsi} MA200={'↑' if above_ma200 else '↓' if above_ma200 is not None else '?'}")

        except Exception as e:
            print(f"  FEL: Kunde inte hämta {ticker_str}: {e}")
            stocks.append({
                "ticker": ticker_str,
                "name": name,
                "price": 0,
                "change": 0,
                "change_pct": 0,
                "high_52w": 0,
                "high_52w_pct": 0,
                "ath": 0,
                "ath_pct": 0,
                "currency": currency,
                "is_alert": False,
                "news": [],
                "domain": resolve_domain(ticker_str),
                "error": str(e),
            })

    return stocks


# Map Yahoo Finance sector → our taxonomy (matches Mac app colors).
YAHOO_SECTOR_MAP = {
    "Technology": "Technology",
    "Financial Services": "Financial",
    "Financial": "Financial",
    "Healthcare": "Healthcare",
    "Consumer Cyclical": "Consumer Discretionary",
    "Consumer Defensive": "Consumer staples",
    "Energy": "Energy",
    "Industrials": "Industrial",
    "Basic Materials": "Industrial",
    "Communication Services": "Technology",
    "Real Estate": "Real Estate",
    "Utilities": "Industrial",
}


def yahoo_sector_industry(ticker: str) -> tuple:
    """Look up (sector, industry) for a ticker via yfinance. Returns ('', '') on failure."""
    try:
        info = yf.Ticker(ticker).info or {}
        raw_sector = (info.get("sector") or "").strip()
        industry = (info.get("industry") or "").strip()
        sector = YAHOO_SECTOR_MAP.get(raw_sector, raw_sector)
        return sector, industry
    except Exception as e:
        print(f"    Yahoo sector lookup failed for {ticker}: {e}")
        return "", ""


_DOMAIN_CACHE: dict | None = None
_DOMAIN_CACHE_DIRTY = False


def _domain_cache_path():
    return SCRIPT_DIR / "domains.json"


def load_domain_cache() -> dict:
    """Persistent cache mapping ticker → company domain (e.g. 'tesla.com').

    Populated on-demand by `resolve_domain()` via yfinance `Ticker.info.website`,
    then saved back to `domains.json` at the end of the run. Avoids hitting
    yfinance `.info` (slow, rate-limited) for tickers we've seen before.
    """
    global _DOMAIN_CACHE
    if _DOMAIN_CACHE is not None:
        return _DOMAIN_CACHE
    path = _domain_cache_path()
    if path.exists():
        try:
            with open(path, "r", encoding="utf-8") as f:
                _DOMAIN_CACHE = json.load(f) or {}
        except Exception as e:
            print(f"  Could not read domains.json: {e}")
            _DOMAIN_CACHE = {}
    else:
        _DOMAIN_CACHE = {}
    return _DOMAIN_CACHE


def save_domain_cache() -> None:
    """Write the domain cache back to disk if anything changed."""
    global _DOMAIN_CACHE, _DOMAIN_CACHE_DIRTY
    if not _DOMAIN_CACHE_DIRTY or _DOMAIN_CACHE is None:
        return
    try:
        path = _domain_cache_path()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(_DOMAIN_CACHE, f, indent=2, sort_keys=True, ensure_ascii=False)
        _DOMAIN_CACHE_DIRTY = False
        print(f"  Saved domains cache ({len(_DOMAIN_CACHE)} entries)")
    except Exception as e:
        print(f"  Could not write domains.json: {e}")


def _normalize_domain(url: str) -> str:
    """Strip scheme, www., trailing slashes/paths from a website URL."""
    if not url:
        return ""
    s = url.strip().lower()
    for prefix in ("https://", "http://"):
        if s.startswith(prefix):
            s = s[len(prefix):]
            break
    if s.startswith("www."):
        s = s[4:]
    # Drop any path/query
    for sep in ("/", "?", "#"):
        i = s.find(sep)
        if i != -1:
            s = s[:i]
    return s.strip()


def resolve_domain(ticker: str) -> str:
    """Return the canonical website domain for a ticker, or '' if unknown.

    Results are cached in `domains.json` so we only hit yfinance `.info` once
    per ticker across runs.
    """
    global _DOMAIN_CACHE, _DOMAIN_CACHE_DIRTY
    cache = load_domain_cache()
    if ticker in cache:
        return cache[ticker] or ""
    domain = ""
    try:
        info = yf.Ticker(ticker).info or {}
        domain = _normalize_domain(info.get("website") or "")
    except Exception as e:
        print(f"    Domain lookup failed for {ticker}: {e}")
    cache[ticker] = domain
    _DOMAIN_CACHE_DIRTY = True
    return domain


def load_stock_meta_overrides() -> dict:
    """Load user-editable sector/industry overrides from stock_meta.json.

    Format:
    {
      "AVGO": {"sector": "Technology", "industry": "Hardware manufacturing"},
      ...
    }
    """
    path = SCRIPT_DIR / "stock_meta.json"
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception as e:
        print(f"  Could not read stock_meta.json: {e}")
        return {}


def load_purchases() -> dict:
    """Load purchases.json (trades per ticker). Returns {} if missing."""
    path = SCRIPT_DIR / "InvestmentApp" / "InvestmentApp" / "purchases.json"
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception as e:
        print(f"  Could not read purchases.json: {e}")
        return {}


def load_splits() -> dict:
    """Load splits.json → {ticker: [(date 'YYYY-MM-DD', ratio float), ...]}.

    Used to reconcile the pre-split trade log with the post-split share count /
    price that Yahoo reports. When a stock splits, add one entry here. Returns
    {} if the file is missing or unreadable (→ no adjustment, same as before).
    """
    path = SCRIPT_DIR / "splits.json"
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f) or {}
    except Exception as e:
        print(f"  Could not read splits.json: {e}")
        return {}
    out = {}
    for tk, events in raw.items():
        if tk.startswith("_") or not isinstance(events, list):
            continue
        parsed = []
        for e in events:
            try:
                d = str(e["date"]); r = float(e["ratio"])
                if r > 0:
                    parsed.append((d, r))
            except (KeyError, TypeError, ValueError):
                continue
        if parsed:
            out[tk.upper()] = parsed
    return out


# Splits Yahoo reports but splits.json doesn't cover yet — surfaced as an in-app
# alert so the user knows to add a splits.json entry (share count/value stay
# wrong until they do). Populated during fetch_all_holdings from data already
# downloaded (no extra Yahoo calls).
_UNHANDLED_SPLITS: list = []

def _flag_unhandled_splits(ticker: str, split_series, handled: dict) -> None:
    """Record any recent Yahoo split for `ticker` not present in splits.json."""
    if split_series is None:
        return
    try:
        events = split_series.dropna()
    except Exception:
        return
    handled_dates = {d for d, _ in handled.get(ticker.upper(), [])}
    try:
        today = datetime.now(timezone.utc).date()
    except Exception:
        return
    for ts, ratio in events.items():
        try:
            r = float(ratio)
        except (TypeError, ValueError):
            continue
        if r in (0.0, 1.0):
            continue
        try:
            d = ts.strftime("%Y-%m-%d")
            age = (today - ts.date()).days
        except Exception:
            continue
        # Only flag splits from the last ~120 days that aren't already handled.
        if 0 <= age <= 120 and d not in handled_dates:
            if not any(u["ticker"] == ticker.upper() and u["date"] == d for u in _UNHANDLED_SPLITS):
                _UNHANDLED_SPLITS.append({"ticker": ticker.upper(), "date": d, "ratio": r})


_FX_CACHE: dict[str, float] = {}

def _currency_to_sek_rate(code: str) -> float:
    """Return `1 unit of <code>` in SEK. Results cached for the process run."""
    code = (code or "SEK").upper()
    if code == "SEK":
        return 1.0
    if code in _FX_CACHE:
        return _FX_CACHE[code]
    try:
        t = yf.Ticker(f"{code}SEK=X")
        hist = t.history(period="5d")
        rate = float(hist["Close"].dropna().iloc[-1]) if not hist.empty else 0.0
    except Exception:
        rate = 0.0
    if rate <= 0:
        rate = {"USD": 9.5, "EUR": 11.0, "DKK": 1.45, "GBP": 12.5,
                "CHF": 11.5, "CAD": 7.0, "NOK": 0.9}.get(code, 1.0)
    _FX_CACHE[code] = rate
    return rate


def merge_purchases_into_holdings(holdings: list) -> list:
    """Ensure every ticker in purchases.json exists in `holdings`, adding new
    ones with sector/industry looked up via Yahoo. Also applies stock_meta.json
    overrides on top of existing holdings.
    """
    purchases = load_purchases()
    meta = load_stock_meta_overrides()
    splits_map = load_splits()

    # Alias map: some tickers in purchases.json are the same stock on a
    # different exchange (e.g. TEAM.MX = Atlassian on BMV, same as TEAM NASDAQ).
    # If the canonical ticker is already in holdings, skip the alias entirely.
    TICKER_ALIASES: dict[str, str] = {
        "TEAM.MX": "TEAM",
        "AZN":     "AZN.ST",
    }

    by_ticker = {h["ticker"]: h for h in holdings}

    # Add new tickers from purchases.json, and update shares/invested for
    # existing tickers so that new purchases are always reflected.
    for ticker, trades in purchases.items():
        if not trades:
            continue
        # Skip if this is an alias for a ticker we already have.
        canonical = TICKER_ALIASES.get(ticker)
        if canonical and canonical in by_ticker:
            continue

        # Average-cost (GAV) accounting, like Avanza: a sell removes shares at the
        # running AVERAGE cost, NOT at the sale price. Subtracting the sale price
        # (the old behaviour) corrupted the cost basis and could drive `invested`
        # negative on profitable sells (e.g. NVDA: bought ~20 USD, sold ~200 USD
        # → invested went negative → bogus -270% return). Process in date order so
        # interleaved buys/sells average correctly.
        # Split adjustment: a trade made BEFORE a split is in pre-split terms.
        # Multiply its shares (and divide its price) by every split ratio that
        # took effect after the trade date, so the running count matches the
        # post-split reality Yahoo prices against. Cost basis (shares*price) is
        # unchanged. Splits are listed in splits.json (see load_splits()).
        ticker_splits = splits_map.get(ticker.upper(), [])

        def _split_factor(trade_date: str) -> float:
            f = 1.0
            for sd, sr in ticker_splits:
                if trade_date and trade_date < sd:
                    f *= sr
            return f

        shares = 0.0
        invested = 0.0
        for t in sorted(trades, key=lambda x: x.get("date") or ""):
            s = float(t.get("shares") or 0)
            p = float(t.get("priceSek") or 0)
            f = _split_factor(t.get("date") or "")
            if f != 1.0:
                s *= f
                p /= f
            if (t.get("type") or "buy") == "sell":
                if shares > 0:
                    avg_cost = invested / shares
                    sold = min(s, shares)        # ignore oversell (data quirk)
                    shares -= sold
                    invested -= sold * avg_cost
            else:
                shares += s
                invested += s * p
        if shares <= 0:
            # Position fully closed (net 0) or oversold. Zero out any EXISTING
            # holding so the downstream `shares > 0` filter drops it. Previously
            # this just `continue`d, which left the stale pre-sale share count in
            # place — so a fully-sold ticker (e.g. ACN) kept showing in the app.
            if ticker in by_ticker:
                by_ticker[ticker]["shares"] = 0
                by_ticker[ticker]["invested"] = 0
            continue
        invested = max(invested, 0.0)            # guard against float drift

        # If ticker already in holdings, just update shares + invested from
        # purchases.json (the authoritative trade log) and move on.
        if ticker in by_ticker:
            h = by_ticker[ticker]
            old_shares = h.get("shares", 0)
            if old_shares != shares:
                print(f"    ↻ Updated shares for {ticker}: {old_shares} → {shares}")
                # Scale the snapshot value proportionally so the price-change
                # calculation in fetch_all_holdings() stays correct.
                if old_shares > 0 and h.get("value", 0):
                    h["value"] = h["value"] * (shares / old_shares)
            h["shares"] = shares
            h["invested"] = round(invested, 2)
            continue

        # Sector/industry: override → Yahoo fallback
        m = meta.get(ticker, {})
        sector = m.get("sector") or ""
        industry = m.get("industry") or ""
        if not sector or not industry:
            ys, yi = yahoo_sector_industry(ticker)
            sector = sector or ys
            industry = industry or yi

        # Guess name from first trade or ticker
        name = ticker
        try:
            info = yf.Ticker(ticker).info or {}
            name = info.get("longName") or info.get("shortName") or ticker
        except Exception:
            pass

        # Guess currency from ticker suffix (fallback to USD)
        if ticker.endswith(".ST"):
            ccy = "SEK"
        elif ticker.endswith(".CO"):
            ccy = "DKK"
        elif ticker.endswith((".MI", ".DE", ".PA", ".AS", ".HE")):
            ccy = "EUR"
        elif ticker.endswith(".L"):
            ccy = "GBP"
        elif ticker.endswith(".SW"):
            ccy = "CHF"
        elif ticker.endswith(".TO"):
            ccy = "CAD"
        else:
            ccy = "USD"

        new_h = {
            "ticker": ticker,
            "name": name,
            "price": 0,
            "shares": shares,
            "value": 0,
            "currency": ccy,
            "sector": sector,
            "industry": industry,
            "pct_of_portfolio": 0,
            "total_return_pct": 0,
            "invested": round(invested, 2),
        }
        holdings.append(new_h)
        by_ticker[ticker] = new_h
        print(f"    + New ticker from purchases.json: {ticker} ({sector} / {industry})")

    # Apply stock_meta overrides to existing holdings
    for ticker, m in meta.items():
        if ticker in by_ticker:
            h = by_ticker[ticker]
            if m.get("sector"):
                h["sector"] = m["sector"]
            if m.get("industry"):
                h["industry"] = m["industry"]

    return holdings


# ---------------------------------------------------------------------------
# FROZEN sector/industry vocabulary (locked 2026-07-08 at the user's request:
# "håll till struktur. om det inte finns någon som passar, blankt"). Any value
# not in these sets is blanked — no new Yahoo category can ever leak in. Only
# extend these lists on the user's explicit instruction. See memory:
# sector-industry-taxonomy.
ALLOWED_SECTORS = {
    "Alternative Assets", "Consumer Discretionary", "Consumer staples", "Energy",
    "Financial", "Healthcare", "Industrial", "Real Estate", "Technology",
}
ALLOWED_INDUSTRIES = {
    "Hardware manufacturing", "Cybersecurity", "Business Efficiency Software",
    "Internet & social media", "Gaming", "IT Services and Consulting", "E-commerce",
    "Media", "Travel and Leisure", "Restaurants & food service", "Online platform",
    "Investment Firms", "Banking", "Insurance", "Payment Services", "Manufacturing",
    "Defence & Aerospace", "Cryptocurrencies", "Precious Metals", "Pharmaceuticals",
    "Healthcare services", "Biotechnology", "Residental", "Commercial", "Personal care",
    "Food and Beverage", "Renewable Energy", "Semiconductors", "Engineering & Construction",
    "Conglomerates", "Financial Data & Stock Exchanges",
}

def snap_taxonomy(value: str, allowed: set) -> str:
    """Return the value if it's in the frozen vocabulary (whitespace/newlines
    collapsed to the canonical form), otherwise blank. Guarantees the SECTOR /
    INDUSTRY columns never show an off-list value."""
    if not value:
        return ""
    norm = " ".join(str(value).split())
    return norm if norm in allowed else ""


def fetch_all_holdings() -> list:
    """Hämta alla holdings från all_holdings.json + live priser via batch yf.download()."""
    config_path = SCRIPT_DIR / "all_holdings.json"
    if not config_path.exists():
        print("  all_holdings.json saknas, hoppar över")
        return []

    with open(config_path, "r", encoding="utf-8") as f:
        holdings = json.load(f)

    # Merge in any new tickers from purchases.json and apply stock_meta.json overrides.
    holdings = merge_purchases_into_holdings(holdings)

    # Filtrera bort ogiltiga entries. (Dropped the old `value < 500000`
    # "rimlighetskoll": the snapshot `value` is scaled by the share ratio on a
    # split — a 4:1 split 4×'d CRWD's snapshot value past 500k and silently
    # dropped it from holdings. The real value_sek is computed live below, so the
    # crude snapshot cap did more harm than good.)
    holdings = [h for h in holdings if h.get("ticker") and h.get("shares", 0) > 0
                and not h["ticker"].startswith("UNKNOWN")]

    tickers = [h["ticker"] for h in holdings]
    if not tickers:
        return []

    # Batch-hämta 1 års data (för ATH 52w)
    try:
        # actions=True adds a "Stock Splits" column per ticker (free — same call)
        # so we can flag Yahoo splits that splits.json doesn't cover yet.
        data = yf.download(tickers, period="1y", group_by="ticker", progress=False, threads=True, actions=True)
    except Exception as e:
        print(f"  FEL vid batch-download: {e}")
        return []

    splits_handled = load_splits()
    results = []
    for h in holdings:
        t = h["ticker"]
        try:
            if len(tickers) == 1:
                closes = data["Close"]
            else:
                closes = data[t]["Close"] if t in data.columns.get_level_values(0) else None

            if closes is None or closes.dropna().empty:
                continue

            # Flag any Yahoo split not yet in splits.json (uses already-downloaded data).
            try:
                if len(tickers) == 1:
                    split_col = data["Stock Splits"] if "Stock Splits" in data.columns else None
                else:
                    sub = data[t] if t in data.columns.get_level_values(0) else None
                    split_col = sub["Stock Splits"] if sub is not None and "Stock Splits" in sub.columns else None
                _flag_unhandled_splits(t, split_col, splits_handled)
            except Exception:
                pass

            clean = closes.dropna()
            price = float(clean.iloc[-1])
            prev = float(clean.iloc[-2]) if len(clean) >= 2 else price
            change_pct = ((price - prev) / prev * 100) if prev else 0
            high_52w = float(clean.max())
            from_ath_pct = ((price - high_52w) / high_52w * 100) if high_52w else 0

            # Beräkna SEK-värde
            shares = h.get("shares", 0)
            invested = h.get("invested", 0)

            # Currency: trust the preset on the holding (set via all_holdings.json
            # or merge_purchases_into_holdings); fall back to Yahoo fast_info,
            # then to a ticker-suffix heuristic as last resort.
            currency = (h.get("currency") or "").upper()
            if not currency:
                try:
                    fi = yf.Ticker(t).fast_info
                    currency = (fi.get("currency") or "").upper() if isinstance(fi, dict) else (getattr(fi, "currency", "") or "").upper()
                except Exception:
                    currency = ""
            if not currency:
                if t.endswith(".ST"):
                    currency = "SEK"
                elif t.endswith(".CO"):
                    currency = "DKK"
                elif t.endswith(".MI") or t.endswith(".DE") or t.endswith(".PA") or t.endswith(".AS"):
                    currency = "EUR"
                elif t.endswith(".L"):
                    currency = "GBP"
                elif t.endswith(".SW"):
                    currency = "CHF"
                elif t.endswith(".HE"):
                    currency = "EUR"
                elif t.endswith(".TO"):
                    currency = "CAD"
                else:
                    currency = "USD"

            # Beräkna SEK-värde direkt från aktuellt pris, antal och valutakurs.
            # För ädelmetaller i gram (unit="g"): pris är per troy oz → dela med 31.1035.
            TROY_OZ_PER_GRAM = 31.1035
            unit = h.get("unit", "")
            fx_rate = _currency_to_sek_rate(currency)
            if unit == "g":
                native_value = (shares / TROY_OZ_PER_GRAM) * price
            else:
                native_value = price * shares
            value_sek = native_value * fx_rate

            # MA200 + RSI(14) — beräknas från redan hämtad 1-årsdata (252 dagar räcker)
            above_ma200 = None
            rsi = None
            try:
                if len(clean) >= 200:
                    ma200_val = clean.rolling(200).mean().iloc[-1]
                    above_ma200 = bool(price > float(ma200_val))
                if len(clean) >= 15:
                    delta = clean.diff()
                    gain = delta.clip(lower=0).rolling(14).mean().iloc[-1]
                    loss = (-delta.clip(upper=0)).rolling(14).mean().iloc[-1]
                    if loss > 0:
                        rsi = round(100 - 100 / (1 + gain / loss), 1)
                    elif gain > 0:
                        rsi = 100.0
                    else:
                        rsi = 50.0
            except Exception as mom_err:
                print(f"    MOM-fel för {t}: {mom_err}")

            ma_str = "↑" if above_ma200 else ("↓" if above_ma200 is not None else "?")
            print(f"    {t}: RSI={rsi} MA200={ma_str} (datapunkter={len(clean)})")

            # Total return recomputed live from current SEK value vs net invested
            # (cost basis). Must NOT be carried forward from the input holding —
            # that field goes stale as soon as price/shares/invested change.
            total_return_pct = round((value_sek - invested) / invested * 100, 1) if invested else 0

            results.append({
                "ticker": t,
                "name": h.get("name", t),
                "price": round(price, 2),
                "change_pct": round(change_pct, 2),
                "shares": shares,
                "invested": round(invested, 2),
                "currency": currency,
                "sector": snap_taxonomy(h.get("sector", ""), ALLOWED_SECTORS),
                "industry": snap_taxonomy(h.get("industry", ""), ALLOWED_INDUSTRIES),
                "pct_of_portfolio": h.get("pct_of_portfolio", 0),
                "total_return_pct": total_return_pct,
                "value_sek": round(value_sek, 0),
                "high_52w": round(high_52w, 2),
                "from_ath_pct": round(from_ath_pct, 1),
                "domain": resolve_domain(t),
                "above_ma200": above_ma200,
                "rsi": rsi,
                **({"unit": unit} if unit else {}),
            })
        except Exception as e:
            print(f"  Hoppar över {t}: {e}")

    save_domain_cache()

    # Recompute pct_of_portfolio live from value_sek — like total_return_pct, it
    # must NOT be carried forward from the input holding. A stale share meant e.g.
    # CRWD (higher value_sek than NVDA after its split) showed a LOWER SHARE %.
    total_value = sum(r.get("value_sek", 0) or 0 for r in results)
    if total_value > 0:
        for r in results:
            r["pct_of_portfolio"] = round((r.get("value_sek", 0) or 0) / total_value * 100, 1)

    results.sort(key=lambda x: x.get("pct_of_portfolio", 0), reverse=True)
    print(f"  Hämtade {len(results)} av {len(holdings)} holdings")
    return results


def fetch_watchlist(held_tickers=None) -> list:
    """Hämta aktier från watchlist.json och bygg samma schema som holdings.

    Skiljer sig från fetch_all_holdings() genom att det inte finns några
    köp/innehav — shares=0, invested=0, value_sek=0 och inget procentuellt
    portföljandel. Syftet är att Swift-appen ska kunna rendera watchlist-
    aktier i samma tabell under en egen sektion.

    `held_tickers`: tickers du redan äger. En bevakad aktie som blivit köpt
    filtreras bort här så att den bara visas under Innehav, inte dubbelt.
    """
    held = {t.upper() for t in (held_tickers or set())}
    # Curated sector/industry overrides — keeps the watchlist on the same
    # vocabulary as holdings instead of raw (granular) Yahoo categories.
    meta_overrides = load_stock_meta_overrides()
    config_path = SCRIPT_DIR / "watchlist.json"
    if not config_path.exists():
        return []

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
    except Exception as e:
        print(f"  watchlist.json kunde inte läsas: {e}")
        return []

    raw = cfg.get("tickers", []) if isinstance(cfg, dict) else cfg
    entries = []
    for item in raw:
        if isinstance(item, str):
            tk = item.upper()
            if tk in held:
                continue
            entries.append({"ticker": tk})
        elif isinstance(item, dict) and item.get("ticker"):
            tk = item["ticker"].upper()
            if tk in held:
                continue
            entries.append({"ticker": tk, **{k: v for k, v in item.items() if k != "ticker"}})
    if held:
        dropped = [ (i if isinstance(i, str) else i.get("ticker","")).upper() for i in raw ]
        dropped = [t for t in dropped if t in held]
        if dropped:
            print(f"  Watchlist: hoppar över {len(dropped)} redan ägda tickers: {', '.join(sorted(set(dropped)))}")
    if not entries:
        return []

    # Watchlist is typically small (1–10 tickers) and the batch yf.download()
    # with group_by="ticker" is flaky for the single-ticker case — it returns
    # different DataFrame shapes depending on yfinance version, which makes
    # `data["Close"]` raise KeyError and silently drop the only entry. Iterate
    # one ticker at a time with yf.Ticker.history() instead, which is both
    # simpler and reliable.
    results = []
    for e in entries:
        t = e["ticker"]
        try:
            tk = yf.Ticker(t)
            hist = tk.history(period="1y", auto_adjust=False)
            if hist is None or hist.empty:
                print(f"  Hoppar över watchlist {t}: ingen prisdata")
                continue
            closes = hist["Close"].dropna()
            if closes.empty:
                print(f"  Hoppar över watchlist {t}: tom Close-serie")
                continue

            price = float(closes.iloc[-1])
            prev = float(closes.iloc[-2]) if len(closes) >= 2 else price
            change_pct = ((price - prev) / prev * 100) if prev else 0
            high_52w = float(closes.max())
            from_ath_pct = ((price - high_52w) / high_52w * 100) if high_52w else 0

            currency = ""
            long_name = t
            sector = ""
            industry = ""
            try:
                try:
                    fi = tk.fast_info
                    currency = (fi.get("currency") or "").upper() if isinstance(fi, dict) else (getattr(fi, "currency", "") or "").upper()
                except Exception:
                    pass
                try:
                    info = tk.info
                    if isinstance(info, dict):
                        long_name = info.get("longName") or info.get("shortName") or t
                        sector = info.get("sector") or ""
                        industry = info.get("industry") or ""
                        if not currency:
                            currency = (info.get("currency") or "").upper()
                except Exception:
                    pass
            except Exception:
                pass

            if not currency:
                if t.endswith(".ST"):
                    currency = "SEK"
                elif t.endswith(".CO"):
                    currency = "DKK"
                elif t.endswith(".MI") or t.endswith(".DE") or t.endswith(".PA") or t.endswith(".AS") or t.endswith(".HE"):
                    currency = "EUR"
                elif t.endswith(".L"):
                    currency = "GBP"
                elif t.endswith(".SW"):
                    currency = "CHF"
                elif t.endswith(".TO"):
                    currency = "CAD"
                else:
                    currency = "USD"

            # Apply curated stock_meta.json overrides (falls back to the Yahoo
            # value when a field isn't overridden; an override of "" blanks it).
            m = meta_overrides.get(t) or meta_overrides.get(t.upper()) or {}
            if "sector" in m:
                sector = m.get("sector") or ""
            if "industry" in m:
                industry = m.get("industry") or ""
            # Enforce the frozen vocabulary (blanks any off-list Yahoo value).
            sector = snap_taxonomy(sector, ALLOWED_SECTORS)
            industry = snap_taxonomy(industry, ALLOWED_INDUSTRIES)

            results.append({
                "ticker": t,
                "name": long_name,
                "price": round(price, 2),
                "change_pct": round(change_pct, 2),
                "shares": 0,
                "invested": 0,
                "currency": currency,
                "sector": sector,
                "industry": industry,
                "pct_of_portfolio": 0,
                "total_return_pct": 0,
                "value_sek": 0,
                "high_52w": round(high_52w, 2),
                "from_ath_pct": round(from_ath_pct, 1),
                "domain": resolve_domain(t),
            })
        except Exception as ex:
            print(f"  Hoppar över watchlist {t}: {ex}")

    save_domain_cache()
    results.sort(key=lambda x: x["ticker"])
    print(f"  Hämtade {len(results)} av {len(entries)} watchlist-aktier")
    return results


def fetch_fx_data() -> dict:
    """Hämta USD/SEK-kursen."""
    try:
        fx = yf.Ticker(FX_TICKER)
        hist = fx.history(period="5d")
        if hist.empty:
            return {"usd_sek": 0, "usd_sek_change_pct": 0}

        current = hist["Close"].iloc[-1]
        prev = hist["Close"].iloc[-2] if len(hist) >= 2 else current
        change_pct = ((current - prev) / prev * 100) if prev else 0

        return {
            "usd_sek": round(float(current), 4),
            "usd_sek_change_pct": round(float(change_pct), 2),
        }
    except Exception as e:
        print(f"  FEL: Kunde inte hämta USD/SEK: {e}")
        return {"usd_sek": 0, "usd_sek_change_pct": 0}


def fetch_commodities() -> dict:
    """Hämta Bitcoin och guldpris i USD."""
    result = {}
    for ticker, name in [("BTC-USD", "bitcoin"), ("GC=F", "gold"), ("CL=F", "oil")]:
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="5d")
            if not hist.empty:
                current = hist["Close"].iloc[-1]
                prev = hist["Close"].iloc[-2] if len(hist) >= 2 else current
                change_pct = ((current - prev) / prev * 100) if prev else 0
                result[name] = {
                    "price": round(float(current), 2),
                    "change_pct": round(float(change_pct), 2),
                }
                print(f"  {name.upper()}: ${float(current):,.2f} ({float(change_pct):+.2f}%)")
            else:
                result[name] = {"price": 0, "change_pct": 0}
        except Exception as e:
            print(f"  FEL: Kunde inte hämta {name}: {e}")
            result[name] = {"price": 0, "change_pct": 0}
    return result


def build_alerts(stocks: list) -> list:
    """Skapa alerts för stora rörelser och andra viktiga händelser."""
    alerts = []

    # Stora rörelser
    for stock in stocks:
        if stock.get("is_alert"):
            direction = "upp" if stock["change_pct"] > 0 else "ner"
            alerts.append({
                "type": "big_move",
                "severity": "high",
                "message": f"{stock['name']} ({stock['ticker']}) rör sig {direction} {abs(stock['change_pct']):.1f}% idag",
                "ticker": stock["ticker"],
            })

    # Nära ATH (inom 5%)
    for stock in stocks:
        if -5 < stock.get("ath_pct", -100) < 0:
            alerts.append({
                "type": "near_ath",
                "severity": "info",
                "message": f"{stock['name']} är {abs(stock['ath_pct']):.1f}% från ATH",
                "ticker": stock["ticker"],
            })

    # Nytt ATH
    for stock in stocks:
        if stock.get("ath_pct", -100) >= 0:
            alerts.append({
                "type": "new_ath",
                "severity": "high",
                "message": f"{stock['name']} har nått nytt ATH!",
                "ticker": stock["ticker"],
            })

    return alerts


# ---------------------------------------------------------------------------
# Nyhetsfilter: bara genuint viktiga nyheter, hellre tomt än brus
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Signalbaserat filter: bara nyheter som påverkar nästa kvartals resultat
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Kända marknadshändelser — injiceras automatiskt baserat på datum
# Format: (start_date, end_date, title_sv, signal_type)
# end_date = None för endagshändelser
# ---------------------------------------------------------------------------
SCHEDULED_EVENTS = [
    # --- 2026 FOMC-möten (Federal Reserve) ---
    ("2026-01-27", "2026-01-28", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-03-17", "2026-03-18", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-05-05", "2026-05-06", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-06-16", "2026-06-17", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-07-28", "2026-07-29", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-09-15", "2026-09-16", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-11-03", "2026-11-04", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    ("2026-12-15", "2026-12-16", "FOMC räntebesked — Federal Reserve möte", "MAKRO"),
    # --- Riksbanken ---
    ("2026-01-29", None, "Riksbanken räntebesked — styrränta 1,75%", "MAKRO"),
    ("2026-03-19", None, "Riksbanken räntebesked + räntebana", "MAKRO"),
    ("2026-05-07", None, "Riksbanken räntebesked", "MAKRO"),
    ("2026-06-17", None, "Riksbanken räntebesked + räntebana", "MAKRO"),
    ("2026-08-20", None, "Riksbanken räntebesked", "MAKRO"),
    ("2026-09-24", None, "Riksbanken räntebesked + räntebana", "MAKRO"),
    ("2026-11-05", None, "Riksbanken räntebesked", "MAKRO"),
    # --- Stora teknikkonferenser ---
    ("2026-03-17", "2026-03-21", "NVIDIA GTC 2026 — Jensen Huang keynote", "AI"),
    ("2026-06-08", "2026-06-12", "Apple WWDC 2026", "KONSUMENT"),
    # --- US makrokalender (Q1 2026) ---
    ("2026-03-28", None, "Core PCE-inflation publiceras (feb)", "MAKRO"),
    ("2026-04-02", None, "ISM Manufacturing PMI (mars)", "MAKRO"),
    ("2026-04-04", None, "Jobbrapport (Nonfarm Payrolls mars)", "MAKRO"),
    ("2026-04-10", None, "CPI-inflation publiceras (mars)", "MAKRO"),
]


def _get_scheduled_events_for_today() -> list:
    """Returnera schemalagda händelser som är aktiva idag, imorgon eller snart."""
    today = datetime.now(TZ_STOCKHOLM).date()
    events = []
    for start_str, end_str, title_sv, signal_type in SCHEDULED_EVENTS:
        start = datetime.strptime(start_str, "%Y-%m-%d").date()
        end = datetime.strptime(end_str, "%Y-%m-%d").date() if end_str else start
        # Show event: 3 days before, during, and day after
        show_start = start - timedelta(days=3)
        show_end = end + timedelta(days=1)
        if show_start <= today <= show_end:
            days_until = (start - today).days
            if days_until > 1:
                label = f"OM {days_until} DAGAR: {title_sv}"
            elif days_until == 1:
                label = f"IMORGON: {title_sv}"
            elif today > end:
                label = title_sv  # day after, still relevant
            elif start == end:
                label = f"IDAG: {title_sv}"
            else:
                label = f"PÅGÅR: {title_sv}"
            events.append({
                "title": label,
                "title_sv": label,
                "publisher": "Kalender",
                "source": "KALENDER",
                "flag": "US" if "riksbank" not in title_sv.lower() else "SE",
                "signal_type": signal_type,
                "impact": "HIGH",
            })
    return events


# De 10 datapunkterna som driver marknaden:
# 1. Fed-räntan  2. Inflation (CPI)  3. Hyperscaler AI-capex
# 4. Nvidia datacenter-intäkter  5. Cloud-tillväxt (AWS/Azure/GCP)
# 6. Digital advertising  7. Apple-efterfrågan  8. Tesla-leveranser
# 9. Jobbdata  10. Likviditet
# + Svenska signaler: Volvo-leveranser, bankräntenetto, Saab ordrar, etc.

IMPORTANT_PATTERNS = [
    # --- Rapporter & resultat (earnings) ---
    r"\bearnings\b", r"\brevenue\b", r"\bprofit\b", r"\bquarterly\b",
    r"\bguidance\b", r"\bforecast\b", r"\boutlook\b",
    r"\bbeats\b", r"\bmisses\b",
    r"\brapport", r"\bvinst", r"\bförlust", r"\bomsättning",
    # --- Fundamentala datapunkter ---
    # AI & datacenter
    r"\bcapex\b", r"\bai spend", r"\bai investment", r"\bdatacenter revenue\b",
    r"\bdata center revenue\b", r"\bgpu\b", r"\bai demand\b",
    r"\bai infrastructure\b", r"\bchip export\b",
    r"\bgtc\b", r"\bkeynote\b", r"\bproduct launch",
    # Cloud-tillväxt
    r"\baws\b.*\bgrowth\b", r"\bazure\b.*\bgrowth\b", r"\bgoogle cloud\b",
    r"\bcloud revenue\b", r"\bcloud growth\b",
    # Digital advertising
    r"\bad revenue\b", r"\bad pricing\b", r"\bad impression",
    r"\badvertising revenue\b", r"\badvertising growth\b",
    # Apple-efterfrågan
    r"\biphone shipment", r"\biphone demand", r"\biphone sale",
    r"\bservices revenue\b", r"\bchina demand\b",
    # Tesla leveranser
    r"\bdeliveries\b", r"\bdelivery\b.*\bnumber", r"\bvehicle delivery",
    r"\bautomotive.{0,10}margin\b",
    # Svenska bolag — fundamentala signaler
    r"\bleverans", r"\border\b", r"\borderingång", r"\borderbok",
    r"\bräntenetto\b", r"\bsubstansvärde\b", r"\bsubstansrabatt",
    r"\bförvaltningsresultat",
    # --- Makro (utökad) ---
    r"\brate cut\b", r"\brate hike\b", r"\bfed\b.*\brate\b", r"\bfederal reserve\b",
    r"\bfomc\b", r"\bpowell\b", r"\bfed meeting\b", r"\bfed decision\b",
    r"\b(?:core )?cpi\b", r"\binflation\b", r"\bstagflation\b",
    r"\bnonfarm payroll", r"\bjobs report\b", r"\bunemployment rate\b",
    r"\bbalance sheet\b.*\bfed\b", r"\bquantitative\b", r"\bliquidity\b",
    r"\briksbank", r"\brecession\b",
    r"\bgdp\b", r"\bpce\b", r"\bcore pce\b", r"\bppi\b",
    r"\bdebt ceiling\b", r"\bgovernment shutdown\b",
    r"\byield curve\b", r"\btreasury\b.*\byield\b", r"\bbond\b.*\byield\b",
    r"\bcredit spread\b", r"\bfinancial stability\b",
    # --- Stora bolagshändelser ---
    r"\bacquisition\b", r"\bacquire", r"\bmerger\b", r"\btakeover\b",
    r"\blayoff", r"\bcut jobs\b", r"\bworkforce reduction\b",
    r"\bceo\b", r"\bcfo\b", r"\bcoo\b", r"\bcto\b", r"\bchief.{0,15}officer\b",
    r"\bresign", r"\bstep.{0,3}down\b", r"\bdepart", r"\breplace",
    r"\bappoint", r"\bnamed\b.*\bceo\b", r"\bnew\b.*\bceo\b",
    r"\binsider\b.*\bsell", r"\binsider\b.*\bbuy", r"\binsider\b.*\btransact",
    r"\bboard\b.*\bmember\b", r"\bstyrelse", r"\bavgår\b", r"\btillträder\b",
    r"\btariff", r"\btrade war\b", r"\bsanction",
    r"\bantitrust\b", r"\bregulat", r"\bsec\b.*\bfil", r"\bipo\b",
    r"\bbuyback\b", r"\bshare repurchase\b", r"\bdividend\b.*\b(?:hike|cut|raise|increase|decrease)\b",
    r"\butdelning",
    r"\bsplit\b", r"\bstock split\b",
    r"\bdowngrade\b", r"\bupgrade\b",
    # --- Geopolitik, risk & sentiment ---
    r"\bmarket crash\b", r"\bsell.?off\b", r"\bcorrection\b",
    r"\bbear market\b", r"\bbull market\b", r"\brally\b",
    r"\bvolatility\b", r"\bvix\b",
    r"\bwar\b", r"\bconflict\b", r"\bgeopolit",
    r"\binvasion\b", r"\bnuclear\b", r"\bmissile\b", r"\bmilitary\b",
    r"\bembargo\b", r"\bfinancial crisis\b", r"\bcontagion\b",
    # --- Energi / råvaror ---
    r"\boil price\b", r"\bcrude\b", r"\bopec\b", r"\bbrent\b", r"\bwti\b",
    r"\bnatural gas\b", r"\blng\b", r"\benergy crisis\b",
    r"\bolja\b", r"\benergi\b",
    r"\bpipeline\b", r"\brefinery\b",
]
_IMPORTANT_RE = re.compile("|".join(IMPORTANT_PATTERNS), re.IGNORECASE)

# Brus som ALLTID filtreras bort
NOISE_KEYWORDS = [
    "best stock", "top stock", "should you buy", "millionaire",
    "weekly picks", "invest $", "rich quick", "don't miss",
    "hot stock", "must-buy", "no-brainer", "smart money move",
    "retire early", "passive income", "dividend king",
    "top pick", "buy now", "wall street loves", "analyst favorite",
    "hidden gem", "under the radar", "oversold", "overbought",
    "3 reasons", "5 reasons", "7 reasons", "10 reasons",
    "you won't believe", "you'll never believe", "here's why",
    "what you need to know", "the next big", "could make you", "get rich",
    "motley fool", "zacks", "tipranks",
    "buy, sell, or hold", "is a buy", "is a sell",
    "have performed since", "follow-up:", "performed since their",
    "etf a buy", "etf to buy", "best etf",
    "since last earnings", "can it continue",
    "up since", "down since", "how has", "how have",
    "pricing in too much", "time to buy", "time to sell",
    "is shifting after", "valuation as p/e", "valuation after",
    "should investors be concerned", "is it too late",
    "should investors", "should you", "could investors",
    "growth stocks to buy", "stocks to buy", "stocks to sell",
    "better risk-adjusted", "risk-adjusted return",
    "just unveiled",  # vague product news
    "tie up", "tie-up",  # partnerships (rarely move earnings)
    "it takes more than", "more than just",
    "what to watch", "things to watch", "keep an eye on",
    "top 2 ", "top 3 ", "top 5 ", "top 10 ",
    # Brus: driftproblem, outages, minor wins
    "outage", "service disruption", "downtime", "went down",
    "experiencing issues", "service issue",
    "lands customer", "wins customer", "signs deal with",
    "lands contract", "small-cap", "penny stock",
    # Bekräftande/status quo — ändrar inte investeringstesen
    "continues to", "remains strong", "still a buy",
    "on track", "reaffirm", "maintains outlook",
    "poised to", "well-positioned", "continues its",
    "doubles down", "bets big on",  # vaga utan siffror
    "what investors should know", "what it means for",
    "why investors", "why this matters",
]


def _is_noise(title: str) -> bool:
    """Filtrera bort brus och clickbait."""
    lower = title.lower()
    return any(kw in lower for kw in NOISE_KEYWORDS)


def _is_important(title: str) -> bool:
    """Kolla om nyheten verkar vara genuint viktig (med ordgränser)."""
    return bool(_IMPORTANT_RE.search(title))


# Alias/nyckelord per bolag — fångar nyheter som inte nämner bolagsnamnet
# exakt men ändå handlar om bolaget (t.ex. "iPhone" → Apple)
COMPANY_ALIASES = {
    # Top holdings
    "NVDA":      ["nvidia", "nvda", "geforce", "rtx", "cuda", "jensen huang", "gtc",
                  "datacenter revenue", "ai chip"],
    "INVE-B.ST": ["investor ab", "investor b", "wallenberg", "substansvärde", "substansrabatt",
                   "patricia industries", "eqt", "atlas copco"],
    "VOLV-B.ST": ["volvo", "volv", "volvo trucks", "volvo leverans"],
    "CRWD":      ["crowdstrike", "crwd"],
    "AAPL":      ["apple", "aapl", "iphone", "ipad", "macbook", "app store",
                  "tim cook", "vision pro", "apple watch", "services revenue"],
    "TSLA":      ["tesla", "tsla", "model y", "model 3", "cybertruck",
                  "elon musk", "musk says", "supercharger", "autopilot",
                  "full self-driving", "vehicle deliveries"],
    "META":      ["meta", "facebook", "instagram", "whatsapp", "zuckerberg",
                  "threads", "meta ai", "llama", "ad revenue"],
    "SEB-A.ST":  ["seb group", "seb bank", "skandinaviska enskilda", "räntenetto"],
    "SHOP":      ["shopify", "shop"],
    "AMZN":      ["amazon", "amzn", "aws", "prime video", "alexa", "andy jassy",
                  "aws growth", "aws revenue"],
    # Focus stocks
    "NFLX":      ["netflix", "nflx", "subscriber"],
    "ZS":        ["zscaler"],
    "GOOGL":     ["google", "alphabet", "googl", "goog", "youtube", "waymo",
                  "google cloud", "azure"],
    "TMDX":      ["transmedics"],
    "SHB-A.ST":  ["handelsbanken", "räntenetto"],
    "SAAB-B.ST": ["saab", "gripen", "orderingång", "orderbok"],
    "IBM":       ["ibm", "red hat", "watsonx"],
    "RBLX":      ["roblox", "rblx"],
    "PLTR":      ["palantir", "pltr"],
    "RKLB":      ["rocket lab", "rklb", "electron", "neutron"],
}


# Kända falska matchningar: US-bolag som delar ticker/namn med svenska bolag
_FALSE_MATCH_BLACKLIST = {
    "INVE-B.ST": ["identiv", "$inve", "royal gold", "rgld"],
    "SEB-A.ST":  ["seaboard", "seb corp"],
    "SHB-A.ST":  ["shb group"],
    "VOLV-B.ST": ["volvo car"],  # Vi trackar Volvo AB (lastbilar), inte Volvo Cars
}


def _mentions_company(title: str, stock: dict) -> bool:
    """Kolla att nyheten PRIMÄRT handlar om bolaget, inte bara nämner det i förbifarten."""
    lower = title.lower()
    ticker = stock["ticker"]

    # Avvisa kända falska matchningar
    for blackword in _FALSE_MATCH_BLACKLIST.get(ticker, []):
        if blackword in lower:
            return False
    name = stock["name"].lower()

    # Exakt ticker eller bolagsnamn i början av rubriken → stark signal
    if lower.startswith(name) or lower.startswith(ticker.lower().replace("-b.st", "").replace("-a.st", "").replace(".st", "")):
        return True

    # Kolla alias-listan
    aliases = COMPANY_ALIASES.get(ticker, [])

    # Stark alias-match: ticker/namn + minst 1 alias → definitivt relevant
    # Svag alias-match: bara 1 generisk alias → kolla att nyheten primärt handlar om bolaget
    # (Undvik: "Applied Materials uses nvidia chips" → taggad som NVDA-nyhet)
    for alias in aliases:
        if alias in lower:
            # Om alias är ett generiskt ord (kort), dubbelkolla att bolaget är subjektet
            if len(alias) <= 4:  # e.g. "meta", "nvda", "ibm", "seb "
                continue  # kräver ytterligare matchning
            return True

    # Korta alias: kräver att bolagsnamn/ticker också finns, ELLER att rubriken börjar med alias
    for alias in aliases:
        if len(alias) <= 4 and alias in lower:
            # Godkänn om rubriken verkar handla om bolaget
            if name in lower or any(a in lower for a in aliases if len(a) > 4):
                return True
            # Godkänn om alias är i första 40 tecken (troligt subjekt)
            pos = lower.find(alias)
            if pos >= 0 and pos < 40:
                return True

    # Fallback: bolagsnamn i rubriken (med ordgräns för att undvika "investerardag" → "investor")
    if name in lower:
        # Kräv ordgräns-matchning för korta generiska namn
        if len(name) <= 8:
            import re as _re
            if _re.search(r'\b' + _re.escape(name) + r'\b', lower):
                return True
        else:
            return True

    return False


def _news_relevance_score(title: str, stock: dict) -> int:
    """Poängsätt en nyhet: högre = mer relevant för bolaget.

    Prioriterar:
    - Bolaget som subjekt (börjar med bolagsnamn) → +50
    - CEO/VD nämns → +30
    - Konkreta siffror (earnings, revenue, $) → +20
    - Produktlansering, förvärv, avtal → +20
    - Bolagsnamn nämns (inte bara partner) → +10
    - Lång rubrik (mer substans) → +5
    - Partner/tredjepartsnyhet (bolaget nämns i förbifarten) → +0
    """
    lower = title.lower()
    name = stock["name"].lower()
    ticker = stock["ticker"].lower().replace("-b.st", "").replace("-a.st", "").replace(".st", "")
    score = 0

    # Bolaget som subjekt (börjar med namn/ticker)
    if lower.startswith(name) or lower.startswith(ticker):
        score += 50

    # Ledningsförändringar (extremt hög signalvärde för långsiktiga investerare)
    leadership_words = ["resign", "step down", "depart", "replace", "appoint",
                        "new ceo", "new cfo", "named ceo", "avgår", "tillträder",
                        "insider sell", "insider buy", "insider transact",
                        "board member", "styrelse"]
    if any(w in lower for w in leadership_words):
        score += 60

    # CEO/VD-uttalanden (hög nyhetsvärde)
    ceo_words = ["ceo", "vd", "chief executive", "jensen huang", "zuckerberg",
                 "musk", "bezos", "cook", "nadella", "pichai", "tobi lütke",
                 "cfo", "coo", "cto", "chief officer"]
    if any(w in lower for w in ceo_words):
        score += 30

    # Konkreta siffror
    import re as _re
    if _re.search(r'\$[\d,.]+[bmBM]|\d+\s*(billion|million|miljarder|miljoner|mdr|%)', lower):
        score += 20

    # Produktlansering, förvärv, avtal
    action_words = ["launches", "acquires", "acquired", "acquisition", "lanserar",
                    "förvärvar", "köper", "partnership", "deal", "contract",
                    "earnings", "revenue", "profit", "guidance", "forecast",
                    "keynote", "announces", "introduced", "unveiled"]
    if any(w in lower for w in action_words):
        score += 20

    # Bolagsnamn nämns explicit (inte bara via kontext)
    if name in lower:
        score += 10

    # Längre rubriker tenderar ha mer substans
    if len(title) > 60:
        score += 5

    return score


def fetch_google_news_rss(query: str, max_items: int = 5) -> list:
    """Hämta nyheter via Google News RSS (ingen API-nyckel behövs)."""
    try:
        import xml.etree.ElementTree as ET
        url = f"https://news.google.com/rss/search?q={requests.utils.quote(query)}&hl=en-US&gl=US&ceid=US:en"
        r = requests.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            root = ET.fromstring(r.content)
            items = []
            for item in root.findall(".//item")[:max_items]:
                title = item.findtext("title", "")
                link = item.findtext("link", "")
                source = item.findtext("source", "")
                if title:
                    items.append({"title": title, "publisher": source, "link": link})
            return items
    except Exception as e:
        print(f"  Google News RSS fel för '{query}': {e}")
    return []


def fetch_finnhub_company_news(ticker: str, days_back: int = 3) -> list:
    """Hämta nyheter för en US-aktie via Finnhub API."""
    api_key = os.environ.get("FINNHUB_API_KEY", "")
    if not api_key:
        return []
    try:
        today = datetime.now(TZ_STOCKHOLM).date()
        from_date = (today - timedelta(days=days_back)).strftime("%Y-%m-%d")
        to_date = today.strftime("%Y-%m-%d")
        r = requests.get(
            f"https://finnhub.io/api/v1/company-news",
            params={"symbol": ticker, "from": from_date, "to": to_date, "token": api_key},
            timeout=10,
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"  Finnhub fel för {ticker}: {e}")
    return []


def collect_top_news(stocks: list, max_items: int = 8) -> list:
    """Samla viktiga bolagsnyheter. Hybrid: Finnhub för US, yfinance för SE.

    Logik:
    - Filtrera bort brus/clickbait (NOISE_KEYWORDS)
    - Nyheten måste nämna bolaget (namn/ticker/alias)
    - Max 1 nyhet per bolag (den mest relevanta)
    """
    seen_titles = set()
    seen_tickers = set()
    important_news = []

    finnhub_key = os.environ.get("FINNHUB_API_KEY", "")

    for stock in stocks:
        ticker = stock["ticker"]
        if ticker in seen_tickers:
            continue

        is_swedish = ticker.endswith(".ST")
        news_items = []

        if finnhub_key and not is_swedish:
            # Finnhub BARA för US-aktier (svenska tickers ger falska träffar
            # t.ex. INVE → Identiv Inc istället för Investor AB)
            fh_news = fetch_finnhub_company_news(ticker)
            for n in fh_news[:20]:  # Scan more items for better relevance
                news_items.append({
                    "title": n.get("headline", ""),
                    "publisher": n.get("source", ""),
                    "link": n.get("url", ""),
                })

        # Komplettera med yfinance (bredare nyhetskällor: Reuters, Bloomberg, etc.)
        yf_news = stock.get("news", [])
        if yf_news:
            for item in yf_news:
                if isinstance(item, dict):
                    content = item.get("content", {})
                    if isinstance(content, dict):
                        title = str(content.get("title", ""))
                        provider = content.get("provider", {})
                        provider_name = str(provider.get("displayName", "")) if isinstance(provider, dict) else ""
                        link = str(content.get("canonicalUrl", {}).get("url", "")) if isinstance(content.get("canonicalUrl"), dict) else ""
                    else:
                        title = str(item.get("title", ""))
                        provider_name = str(item.get("publisher", ""))
                        link = str(item.get("link", ""))
                    if title and title not in {n.get("title", "") for n in news_items}:
                        news_items.append({
                            "title": title,
                            "publisher": provider_name,
                            "link": link,
                        })

        # Google News RSS som tredje källa (fångar Bloomberg, CNBC, Reuters)
        if not news_items or len(news_items) < 5:
            search_name = stock["name"]
            if is_swedish:
                # Sök med fullständigt bolagsnamn för svenska aktier
                # (undvik "SEB stock" → Seaboard Corp)
                search_name = stock["name"]
                google_news = fetch_google_news_rss(f'"{search_name}" aktie OR stock', max_items=5)
            else:
                google_news = fetch_google_news_rss(f"{search_name} stock", max_items=5)
            existing_titles = {n.get("title", "") for n in news_items}
            for gn in google_news:
                if gn["title"] not in existing_titles:
                    news_items.append(gn)

        # Samla alla kandidater och välj den bästa (inte bara första)
        candidates = []
        for item in news_items:
            title = item.get("title", "") or item.get("headline", "")
            if not title or title in seen_titles:
                continue
            if _is_noise(title):
                continue
            if not _mentions_company(title, stock):
                continue
            score = _news_relevance_score(title, stock)
            candidates.append((score, item, title))

        if candidates:
            # Välj nyheten med högst relevanspoäng
            candidates.sort(key=lambda x: x[0], reverse=True)
            best_score, best_item, best_title = candidates[0]
            seen_titles.add(best_title)
            seen_tickers.add(ticker)
            important_news.append({
                **best_item,
                "title": best_title,
                "ticker": ticker,
                "stock_name": stock["name"],
                "signal_type": _classify_signal(best_title),
                "impact": _score_to_impact(best_score),
            })

    return important_news[:max_items]


# Persistenta cachefiler för nyhetssammanfattningar.
# Throttle: kör Claude max 1 gång per 4 timmar. Mellan körningar återanvänds
# title_sv från cachen för titlar vi redan sett. Sparar ~90 % av token-kostnaden
# eftersom cron pingas var 15-30:e min men nyheter sällan ändras så snabbt.
_NEWS_CACHE_PATH = SCRIPT_DIR / "news_summary_cache.json"
_NEWS_THROTTLE_HOURS = 4


def _load_news_cache() -> dict:
    """Returnerar {title_hash: title_sv} + last_run_iso. Tomma defaults om saknas."""
    if not _NEWS_CACHE_PATH.exists():
        return {"last_run": None, "summaries": {}}
    try:
        with open(_NEWS_CACHE_PATH, "r") as f:
            data = json.load(f)
        return {"last_run": data.get("last_run"), "summaries": data.get("summaries", {})}
    except Exception as e:
        print(f"  News-cache läsfel ({e}) — börjar om")
        return {"last_run": None, "summaries": {}}


def _save_news_cache(cache: dict) -> None:
    try:
        with open(_NEWS_CACHE_PATH, "w") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  Kunde inte spara news-cache: {e}")


def _title_key(title: str) -> str:
    """Stabil nyckel — case-insensitive, trimmar whitespace."""
    return title.strip().lower()


def summarize_news(company_news: list, market_news: list) -> tuple[list, list]:
    """Använd Claude för att skriva korta, tydliga svenska sammanfattningar.

    Throttle: vi anropar Claude max en gång per _NEWS_THROTTLE_HOURS timmar.
    Mellan körningar fylls title_sv på från cachen (matchar på title-hash).
    Nyheter som varken finns i cachen eller hinner med Claude-rundan får
    ingen svensk text (frontend faller tillbaka på original).
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("  Ingen ANTHROPIC_API_KEY — visar originalrubriker")
        return company_news, market_news

    # 1. Ladda cache och fyll på title_sv för redan sedda titlar
    cache = _load_news_cache()
    cached_summaries: dict = cache["summaries"]
    for n in company_news + market_news:
        sv = cached_summaries.get(_title_key(n["title"]))
        if sv:
            n["title_sv"] = sv

    # 2. Throttle-check: har vi kört Claude nyligen?
    now = datetime.now(timezone.utc)
    last_run_iso = cache.get("last_run")
    if last_run_iso:
        try:
            last_run = datetime.fromisoformat(last_run_iso)
            age_hours = (now - last_run).total_seconds() / 3600
            if age_hours < _NEWS_THROTTLE_HOURS:
                cached_count = sum(1 for n in company_news + market_news if n.get("title_sv"))
                total = len(company_news) + len(market_news)
                print(
                    f"  Claude throttle: {age_hours:.1f}h sedan senaste körning "
                    f"(<{_NEWS_THROTTLE_HOURS}h) — använder cache "
                    f"({cached_count}/{total} träffar)"
                )
                return company_news, market_news
        except Exception:
            pass  # Trasig timestamp → kör ändå

    # 3. Bygg lista till Claude (alla titlar — billigare att skicka alla än att
    # bygga separat prompt för bara nya. ~16 titlar är försumbart.)
    all_items = []
    for n in company_news:
        short_ticker = n["ticker"].replace("-B.ST", "").replace("-A.ST", "").replace(".ST", "")
        all_items.append(f"[{short_ticker}] {n['title']}")
    for n in market_news:
        all_items.append(f"[{n['source']}] {n['title']}")

    if not all_items:
        return company_news, market_news

    numbered = "\n".join(f"{i+1}. {t}" for i, t in enumerate(all_items))

    try:
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1200,
            messages=[{
                "role": "user",
                "content": f"""Du är en nyhetsredaktör för en LÅNGSIKTIG investerare (3-5 år horisont). Bara nyheter som ÄNDRAR investeringstesen.

BEHÅLL — ändrar narrativet:
- Kvartalsrapporter, intäkter, vinst, guidance-ändringar
- VD/CFO avgår, ny VD, stora insiderförsäljningar
- Förvärv, avknoppningar, stora avtal med siffror
- Nya produktkategorier, materiella lanseringar
- Makro: Fed-beslut, CPI, jobbdata, räntebesked
- Regulatoriska beslut, antitrust, tullar
- Stora uppsägningar, omstruktureringar
- Utdelningsändringar, aktiesplittar

FILTRERA BORT (SKIP):
- Bekräftande nyheter ("Meta satsar på AI" — alla vet)
- Prisrörelser utan orsak ("stock up 3%")
- Småpartnerskap utan siffror
- Analytikerklickbait, listicles
- Historisk avkastning
- Duplicerade nyheter (behåll bästa)

VIKTIGT: Behåll hellre en nyhet för mycket än en för lite!

REGLER:
- Skriv om BEHÅLLNA nyheter till korta, tydliga svenska meningar
- STRIKT MAX 55 TECKEN per mening! Räkna tecknen. Hugga ALDRIG av med "..."
- Korta ner: "miljarder" → "mdr", "dollar" → "$", skippa onödiga ord
- Var konkret: vad hände? Siffror om möjligt
- Nyheter som filtreras bort: skriv numret följt av "SKIP"
- Svara BARA med numrerade rader

{numbered}"""
            }],
        )

        lines = msg.content[0].text.strip().split("\n")
        summaries = {}
        skipped = set()
        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Parsa "1. sammanfattning" eller "1. SKIP"
            parts = line.split(".", 1)
            if len(parts) == 2 and parts[0].strip().isdigit():
                idx = int(parts[0].strip()) - 1
                text = parts[1].strip()
                if text.upper() == "SKIP":
                    skipped.add(idx)
                else:
                    # Säkerhetsnät: trunkera till 65 tecken om Claude inte följer gränsen
                    if len(text) > 65:
                        # Hitta sista mellanslag före 62 tecken
                        cut = text[:62].rfind(" ")
                        if cut > 25:
                            text = text[:cut]
                            # Ta bort hängande prepositioner/konjunktioner
                            for suffix in [" för", " i", " på", " av", " med", " efter", " till", " om"]:
                                if text.endswith(suffix):
                                    text = text[:-len(suffix)]
                                    break
                            text += "."
                    summaries[idx] = text

        # Uppdatera company_news — ta bort SKIP:ade
        for i, n in enumerate(company_news):
            if i in summaries:
                n["title_sv"] = summaries[i]
            elif i in skipped:
                n["_skip"] = True

        # Uppdatera market_news (offset = len(company_news))
        offset = len(company_news)
        for i, n in enumerate(market_news):
            if (offset + i) in summaries:
                n["title_sv"] = summaries[offset + i]
            elif (offset + i) in skipped:
                n["_skip"] = True

        # Filtrera bort SKIP:ade nyheter
        company_news = [n for n in company_news if not n.get("_skip")]
        market_news = [n for n in market_news if not n.get("_skip")]

        kept = len(summaries)
        total = kept + len(skipped)
        print(f"  Claude: behöll {kept}/{total} nyheter, filtrerade bort {len(skipped)} brus")

        # Uppdatera cache: spara alla nya title_sv keyade på title-hash, plus
        # behåll de gamla (limit 200 entries så filen inte växer obegränsat).
        new_entries = {}
        for n in company_news + market_news:
            sv = n.get("title_sv")
            if sv:
                new_entries[_title_key(n["title"])] = sv
        merged = {**cached_summaries, **new_entries}
        # Trim till de 200 senast använda (insertion order = senast tillagda sist)
        if len(merged) > 200:
            merged = dict(list(merged.items())[-200:])
        _save_news_cache({"last_run": now.isoformat(), "summaries": merged})

    except Exception as e:
        print(f"  FEL vid AI-sammanfattning: {e}")

    return company_news, market_news


def fetch_upcoming_earnings(holdings: list, days_ahead: int = 7) -> list:
    """Kolla vilka bolag som rapporterar inom de närmaste dagarna."""
    upcoming = []
    now = datetime.now(TZ_STOCKHOLM)
    cutoff = now + timedelta(days=days_ahead)

    for holding in holdings:
        try:
            t = yf.Ticker(holding["ticker"])
            cal = t.calendar
            if cal is None:
                continue

            earnings_date = None

            # Dict-format (nyare yfinance)
            if isinstance(cal, dict):
                ed_val = cal.get("Earnings Date")
                if isinstance(ed_val, list) and len(ed_val) > 0:
                    earnings_date = ed_val[0]
                elif hasattr(ed_val, "date"):
                    earnings_date = ed_val

            # DataFrame-format (äldre yfinance)
            elif hasattr(cal, "iloc"):
                try:
                    cols = list(cal.columns) if hasattr(cal, "columns") else []
                    idx = list(cal.index) if hasattr(cal, "index") else []
                    if "Earnings Date" in cols:
                        earnings_date = cal["Earnings Date"].iloc[0]
                    elif "Earnings Date" in idx:
                        earnings_date = cal.loc["Earnings Date"].iloc[0]
                except Exception:
                    pass

            if earnings_date is not None:
                from datetime import date as _date_type
                if isinstance(earnings_date, _date_type):
                    ed_date = earnings_date
                elif hasattr(earnings_date, "date"):
                    ed_date = earnings_date.date()
                elif isinstance(earnings_date, str):
                    ed_date = datetime.strptime(earnings_date[:10], "%Y-%m-%d").date()
                else:
                    continue

                now_date = now.date()
                cutoff_date = cutoff.date()

                if now_date <= ed_date <= cutoff_date:
                    days_until = (ed_date - now_date).days
                    day_name = WEEKDAY_SV[ed_date.weekday()]
                    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
                    date_display = f"{ed_date.day} {month_names[ed_date.month - 1]}"

                    upcoming.append({
                        "ticker": holding["ticker"],
                        "name": holding["name"],
                        "date": ed_date.strftime("%Y-%m-%d"),
                        "date_display": date_display,
                        "days_until": days_until,
                    })
        except Exception:
            pass

    upcoming.sort(key=lambda x: x["date"])
    return upcoming


def fetch_upcoming_dividends(holdings: list, days_ahead: int = 90) -> list:
    """Hämta kommande utdelningar (ex-datum, belopp, yield) för alla innehav."""
    upcoming = []
    now = datetime.now(TZ_STOCKHOLM)
    now_date = now.date()
    cutoff_date = (now + timedelta(days=days_ahead)).date()
    month_names = ["jan", "feb", "mar", "apr", "maj", "jun",
                   "jul", "aug", "sep", "okt", "nov", "dec"]

    for holding in holdings:
        try:
            t = yf.Ticker(holding["ticker"])
            cal = t.calendar
            if cal is None:
                continue

            ex_date_raw = None

            # Dict-format (nyare yfinance)
            if isinstance(cal, dict):
                ed_val = cal.get("Ex-Dividend Date")
                if isinstance(ed_val, list) and len(ed_val) > 0:
                    ex_date_raw = ed_val[0]
                elif ed_val is not None:
                    ex_date_raw = ed_val

            # DataFrame-format (äldre yfinance)
            elif hasattr(cal, "iloc"):
                try:
                    cols = list(cal.columns) if hasattr(cal, "columns") else []
                    idx = list(cal.index) if hasattr(cal, "index") else []
                    if "Ex-Dividend Date" in cols:
                        ex_date_raw = cal["Ex-Dividend Date"].iloc[0]
                    elif "Ex-Dividend Date" in idx:
                        ex_date_raw = cal.loc["Ex-Dividend Date"].iloc[0]
                except Exception:
                    pass

            if ex_date_raw is None:
                continue

            # Konvertera till date
            from datetime import date as _date_type
            if isinstance(ex_date_raw, _date_type):
                ex_date = ex_date_raw
            elif hasattr(ex_date_raw, "date"):
                ex_date = ex_date_raw.date()
            elif isinstance(ex_date_raw, (int, float)):
                ex_date = datetime.fromtimestamp(ex_date_raw).date()
            elif isinstance(ex_date_raw, str):
                ex_date = datetime.strptime(ex_date_raw[:10], "%Y-%m-%d").date()
            else:
                continue

            if ex_date < now_date or ex_date > cutoff_date:
                continue

            days_until = (ex_date - now_date).days
            date_display = f"{ex_date.day} {month_names[ex_date.month - 1]}"

            # Hämta utdelningsbelopp och yield
            info = t.info or {}
            amount = info.get("lastDividendValue")
            yield_pct = info.get("dividendYield")
            if yield_pct:
                # yfinance returnerar yield redan i procent (4.62 = 4.62%)
                yield_pct = round(float(yield_pct), 2)

            currency = "SEK" if holding["ticker"].endswith(".ST") else "USD"

            upcoming.append({
                "ticker": holding["ticker"],
                "name": holding["name"],
                "ex_date": ex_date.strftime("%Y-%m-%d"),
                "ex_date_display": date_display,
                "days_until": days_until,
                "amount": round(float(amount), 2) if amount else None,
                "currency": currency,
                "yield_pct": round(float(yield_pct), 2) if yield_pct else None,
            })
        except Exception as e:
            print(f"  Dividend-fel för {holding['ticker']}: {e}")

    upcoming.sort(key=lambda x: x["ex_date"])
    return upcoming


def fetch_portfolio_value() -> dict:
    """Hämta totalt portföljvärde från investments.xlsx via Dropbox."""
    xlsx_url = os.environ.get("DROPBOX_XLSX_URL")
    if not xlsx_url:
        # Lokalt läge — läs direkt från disk
        local_path = Path.home() / "Library/CloudStorage/Dropbox/investments.xlsx"
        if local_path.exists():
            print(f"  Läser lokal fil: {local_path}")
            wb = openpyxl.load_workbook(local_path, data_only=True, read_only=True)
            ws = wb["Investments"]
            value = None
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=47, max_col=47, values_only=True):
                value = row[0]
            wb.close()
            if value and isinstance(value, (int, float)):
                print(f"  Portföljvärde: {value:,.0f} kr")
                return {"value_sek": round(float(value))}
        print("  Ingen DROPBOX_XLSX_URL och ingen lokal fil — hoppar över portfölj")
        return {}

    try:
        print(f"  Laddar ner xlsx från Dropbox...")
        r = requests.get(xlsx_url, timeout=90)
        if r.status_code != 200:
            print(f"  FEL: Dropbox svarade {r.status_code}")
            return {}

        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
        ws = wb["Investments"]

        # AU1 = kolumn 47 = totalt portföljvärde
        value = None
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=47, max_col=47, values_only=True):
            value = row[0]
        wb.close()

        if value and isinstance(value, (int, float)):
            print(f"  Portföljvärde: {value:,.0f} kr")
            return {"value_sek": round(float(value))}
        else:
            print(f"  VARNING: AU1 = {value} (oväntat format)")
            return {}

    except Exception as e:
        print(f"  FEL: Kunde inte hämta portföljvärde: {e}")
        return {}


def fetch_previous_portfolio() -> dict:
    """Hämta föregående portföljvärde från befintlig Gist-data.

    Hanterar helger korrekt: sparar fredagens stängningsvärde
    som prev_close så att helg-körningar visar riktig förändring.
    """
    try:
        old_data = _read_private_market_data()
        if old_data:
            port = old_data.get("portfolio", {})
            old_value = port.get("value_sek", 0)
            old_date = old_data.get("updated_at", "")[:10]
            today = datetime.now(TZ_STOCKHOLM)
            today_str = today.strftime("%Y-%m-%d")
            weekday = today.weekday()  # 0=mån, 5=lör, 6=sön

            # Bevara portfolio_history från tidigare data
            old_history = port.get("history", [])

            result = {
                "history": old_history,
            }

            if old_date == today_str:
                # Samma dag — behåll befintlig prev_close
                result["prev_close_sek"] = port.get("prev_close_sek") or old_value
            elif weekday in (5, 6):
                # Helg — använd fredagens prev_close (som redan var satt)
                # så att vi inte nollställer mot helg-värdet
                result["prev_close_sek"] = port.get("prev_close_sek") or old_value
            else:
                # Ny börsdag — gårdagens sista värde blir prev_close
                result["prev_close_sek"] = old_value

            return result
    except Exception as e:
        print(f"  Kunde inte hämta föregående portföljvärde: {e}")
    return {}


def fetch_fear_greed() -> dict:
    """Hämta CNN Fear & Greed Index."""
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
        r = requests.get(
            "https://production.dataviz.cnn.io/index/fearandgreed/graphdata",
            headers=headers,
            timeout=10,
        )
        if r.status_code == 200:
            d = r.json()
            fg = d.get("fear_and_greed", {})
            score = round(fg.get("score", 0))
            rating = fg.get("rating", "neutral")
            print(f"  Fear & Greed: {score} ({rating})")
            return {"score": score, "rating": rating}
    except Exception as e:
        print(f"  FEL: Kunde inte hämta Fear & Greed: {e}")
    return {"score": 0, "rating": "neutral"}


def build_weekly_summary(stocks: list, portfolio: dict, indices: dict) -> dict:
    """Bygg en veckosammanfattning för helgvisning."""
    now = datetime.now(TZ_STOCKHOLM)
    if now.weekday() not in (5, 6):  # Bara helg
        return {}

    # Beräkna veckoförändring per aktie (5d data)
    weekly_changes = []
    for stock in stocks:
        weekly_changes.append({
            "name": stock["name"],
            "ticker": stock["ticker"],
            "change_pct": stock.get("change_pct", 0),
            "price": stock.get("price", 0),
        })

    weekly_changes.sort(key=lambda s: s["change_pct"], reverse=True)
    best = weekly_changes[0] if weekly_changes else None
    worst = weekly_changes[-1] if weekly_changes else None

    # Portföljförändring (vecka: mån → fre via historik)
    port_val = portfolio.get("value_sek", 0)
    history = portfolio.get("history", [])

    # Hitta veckans start (måndag) och slut (fredag) i historiken
    today = now.date()
    # Beräkna måndagens datum (sön: -6, lör: -5)
    days_since_monday = today.weekday()  # lör=5, sön=6
    monday = today - timedelta(days=days_since_monday)
    monday_str = monday.strftime("%Y-%m-%d")
    friday_str = (monday + timedelta(days=4)).strftime("%Y-%m-%d")

    week_start_val = None
    week_end_val = None
    for h in history:
        d = h.get("date", "")
        if d == monday_str:
            week_start_val = h.get("value")
        if d == friday_str:
            week_end_val = h.get("value")

    # Fallback: använd äldsta/nyaste i veckans range
    if not week_start_val or not week_end_val:
        week_entries = [h for h in history if monday_str <= h.get("date", "") <= friday_str]
        if len(week_entries) >= 2:
            week_start_val = week_start_val or week_entries[0].get("value")
            week_end_val = week_end_val or week_entries[-1].get("value")

    summary_parts = []
    if port_val and week_start_val and week_end_val:
        week_change = week_end_val - week_start_val
        week_pct = (week_change / week_start_val * 100) if week_start_val else 0
        direction = "upp" if week_change >= 0 else "ner"
        summary_parts.append(
            f"Portföljen {direction} {abs(week_pct):.1f}% denna vecka ({week_change:+,.0f} kr). Värde: {port_val:,.0f} kr."
        )
    elif port_val:
        summary_parts.append(f"Portföljen stängde på {port_val:,.0f} kr.")

    if best and best["change_pct"] > 0:
        summary_parts.append(f"Veckans vinnare: {best['name']} +{best['change_pct']:.1f}%.")
    if worst and worst["change_pct"] < 0:
        summary_parts.append(f"Veckans förlorare: {worst['name']} {worst['change_pct']:.1f}%.")

    sp = indices.get("sp500", {})
    omx = indices.get("omxs30", {})
    if sp.get("change_pct"):
        summary_parts.append(f"S&P 500: {sp['change_pct']:+.1f}%.")
    if omx.get("change_pct"):
        summary_parts.append(f"OMXS30: {omx['change_pct']:+.1f}%.")

    return {
        "text": " ".join(summary_parts),
        "best": best,
        "worst": worst,
    }


def fetch_indices() -> dict:
    """Hämta daglig förändring för S&P 500 och OMXS30."""
    result = {}
    for ticker, name in [("^GSPC", "sp500"), ("^OMX", "omxs30")]:
        try:
            t = yf.Ticker(ticker)
            hist = t.history(period="5d")
            if not hist.empty:
                current = hist["Close"].iloc[-1]
                prev = hist["Close"].iloc[-2] if len(hist) >= 2 else current
                change_pct = ((current - prev) / prev * 100) if prev else 0
                result[name] = {
                    "price": round(float(current), 2),
                    "change_pct": round(float(change_pct), 2),
                }
                print(f"  {name.upper()}: {float(current):,.0f} ({float(change_pct):+.2f}%)")
            else:
                result[name] = {"price": 0, "change_pct": 0}
        except Exception as e:
            print(f"  FEL: Kunde inte hämta {name}: {e}")
            result[name] = {"price": 0, "change_pct": 0}
    return result


# Makro-nyckelord: nyheter som matchar dessa klassas som MAKRO
_MACRO_RE = re.compile(
    r"\b(?:federal reserve|fomc|powell|rate cut|rate hike|fed rate|fed meeting|fed decision|"
    r"cpi|inflation|core cpi|consumer price|stagflation|"
    r"gdp|pce|core pce|ppi|"
    r"nonfarm payroll|jobs report|unemployment rate|jobless claim|"
    r"balance sheet|quantitative|liquidity|financial condition|"
    r"riksbank|ecb|central bank|"
    r"tariff|trade war|recession|sanction|"
    r"treasury yield|bond yield|yield curve|credit spread|"
    r"debt ceiling|government shutdown|"
    r"vix|volatility|sell.?off|correction|bear market|bull market|rally|"
    r"geopolit|war|conflict)\b",
    re.IGNORECASE,
)


def _score_to_impact(score: int) -> str:
    """Mappa relevanspoäng till impaktnivå för visuell prioritering."""
    if score >= 80:
        return "HIGH"
    elif score >= 40:
        return "MED"
    return "LOW"


def _classify_signal(title: str) -> str:
    """Klassificera en nyhet: ENERGI, RISK, MAKRO, AI, CLOUD, KONSUMENT, LEDNING eller SIGNAL."""
    lower = title.lower()
    # Geopolitisk risk / systemrisk (högst prioritet — påverkar hela marknaden)
    if any(kw in lower for kw in ["war", "conflict", "invasion", "sanction", "geopolit",
                                    "nuclear", "missile", "military", "nato",
                                    "trade war", "tariff", "embargo",
                                    "debt ceiling", "government shutdown", "default",
                                    "financial crisis", "systemic risk", "contagion",
                                    "krig", "konflikt", "sanktion"]):
        return "RISK"
    # Energi / råvaror (driver inflation, valutor, hela sektorer)
    if any(kw in lower for kw in ["oil", "crude", "opec", "brent", "wti",
                                    "natural gas", "lng", "energy crisis",
                                    "olja", "energi", "energikris",
                                    "oil price", "oil surge", "oil spike",
                                    "pipeline", "refinery", "petroleum"]):
        return "ENERGI"
    # Ledningsförändringar
    if any(kw in lower for kw in ["ceo", "cfo", "resign", "step down", "depart",
                                    "appoint", "avgår", "tillträder", "insider sell",
                                    "insider buy", "board member", "styrelse"]):
        return "LEDNING"
    if _MACRO_RE.search(title):
        return "MAKRO"
    if any(kw in lower for kw in ["capex", "ai spend", "ai invest", "datacenter revenue",
                                    "data center revenue", "gpu", "ai demand", "nvidia",
                                    "ai infrastructure"]):
        return "AI"
    if any(kw in lower for kw in ["aws", "azure", "google cloud", "cloud revenue",
                                    "cloud growth"]):
        return "CLOUD"
    if any(kw in lower for kw in ["iphone", "apple", "tesla", "deliveries", "delivery",
                                    "ad revenue", "advertising"]):
        return "KONSUMENT"
    return "SIGNAL"


def fetch_market_news(max_items: int = 8) -> list:
    """Hämta viktiga övergripande marknadsnyheter.

    Hybrid: Finnhub general news (bättre kvalitet) med yfinance som fallback.
    Fokus: makro-signaler (Fed, CPI, jobb, likviditet) + hyperscaler-signaler.
    """
    all_news = []
    finnhub_key = os.environ.get("FINNHUB_API_KEY", "")

    if finnhub_key:
        # Finnhub general news — Reuters, CNBC, etc.
        try:
            r = requests.get(
                "https://finnhub.io/api/v1/news",
                params={"category": "general", "token": finnhub_key},
                timeout=10,
            )
            if r.status_code == 200:
                for item in r.json()[:80]:  # Scan more items
                    title = item.get("headline", "")
                    source = item.get("source", "MKT")
                    if title and not _is_noise(title) and _is_important(title):
                        signal_type = _classify_signal(title)
                        # Marknadsnyheter: MAKRO=HIGH, AI/CLOUD=MED, annat=LOW
                        impact = "HIGH" if signal_type in ("MAKRO", "RISK", "ENERGI") else ("MED" if signal_type in ("AI", "CLOUD", "LEDNING") else "LOW")
                        all_news.append({
                            "title": title,
                            "publisher": source,
                            "source": source,
                            "flag": "US",
                            "signal_type": signal_type,
                            "impact": impact,
                        })
        except Exception as e:
            print(f"  Finnhub marknadsnyheter fel: {e}")

    if not all_news:
        # Fallback: yfinance
        for ticker_sym, label, flag in [("^GSPC", "S&P", "US"), ("^OMX", "OMX", "SE")]:
            try:
                t = yf.Ticker(ticker_sym)
                ticker_news = t.news or []
                for item in ticker_news[:5]:
                    if isinstance(item, dict):
                        content = item.get("content", {})
                        if isinstance(content, dict):
                            title = str(content.get("title", ""))
                            provider = content.get("provider", {})
                            provider_name = str(provider.get("displayName", "")) if isinstance(provider, dict) else ""
                        else:
                            title = str(item.get("title", ""))
                            provider_name = str(item.get("publisher", ""))
                        if title and not _is_noise(title) and _is_important(title):
                            signal_type = _classify_signal(title)
                            impact = "HIGH" if signal_type in ("MAKRO", "RISK", "ENERGI") else ("MED" if signal_type in ("AI", "CLOUD", "LEDNING") else "LOW")
                            all_news.append({
                                "title": title,
                                "publisher": provider_name,
                                "source": label,
                                "flag": flag,
                                "signal_type": signal_type,
                                "impact": impact,
                            })
            except Exception as e:
                print(f"  Kunde inte hämta marknadsnyheter för {ticker_sym}: {e}")

    # Prioritera: MAKRO först, sedan AI, CLOUD, KONSUMENT
    priority = {"MAKRO": 0, "AI": 1, "CLOUD": 2, "KONSUMENT": 3, "SIGNAL": 4}
    all_news.sort(key=lambda n: priority.get(n.get("signal_type", "SIGNAL"), 4))

    # Deduplicera + begränsa geopolitik/krig till max 2 rubriker
    seen = set()
    unique = []
    geo_count = 0
    geo_keywords = ["war", "iran", "conflict", "troops", "soldiers", "military", "drone", "strait of hormuz"]
    for item in all_news:
        if item["title"] not in seen:
            seen.add(item["title"])
            lower_title = item["title"].lower()
            is_geo = any(kw in lower_title for kw in geo_keywords)
            if is_geo:
                geo_count += 1
                if geo_count > 2:
                    continue  # Max 2 geopolitik-rubriker
            unique.append(item)

    return unique[:max_items]


# ---------------------------------------------------------------------------
# Gist-publicering
# ---------------------------------------------------------------------------

class SafeEncoder(json.JSONEncoder):
    """JSON-encoder som hanterar numpy-typer och frozendicts."""

    def default(self, obj):
        # Hantera numpy-typer
        if hasattr(obj, "item"):
            return obj.item()
        # Hantera frozendict
        if hasattr(obj, "items") and type(obj).__name__ == "frozendict":
            return dict(obj)
        return super().default(obj)


def _sanitize_nan(obj):
    """Ersätt NaN/Infinity rekursivt med None så att JSON alltid är giltig."""
    import math
    if isinstance(obj, float):
        return None if (math.isnan(obj) or math.isinf(obj)) else obj
    if isinstance(obj, dict):
        return {k: _sanitize_nan(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_sanitize_nan(v) for v in obj]
    return obj


def _safe_json(data, **kwargs):
    """Serialisera data till JSON med SafeEncoder, NaN → null."""
    return json.dumps(_sanitize_nan(data), cls=SafeEncoder, ensure_ascii=False, **kwargs)


COMPANY_INFO_MAX_AGE_DAYS = 7        # refresh a ticker's fundamentals at most weekly
COMPANY_INFO_MAX_FETCH_PER_RUN = 8   # cap slow .info calls per run (cold start spreads out)


def _trim_sentences(text: str, n: int = 3) -> str:
    text = " ".join((text or "").split())
    if not text:
        return ""
    parts = re.split(r"(?<=[.!?])\s+", text)
    return " ".join(parts[:n]).strip()


def build_company_info(tickers: list, previous: dict, today: str) -> dict:
    """Description + 6 key metrics per ticker for the app's detail view.

    Cached in the gist itself (the workflow commits no files back, so the
    previous gist is our only durable store): carry `previous` forward and only
    hit the slow yfinance `.info` for tickers that are missing or older than a
    week, capped per run so a cold start spreads across several runs.
    """
    out = {t: dict(v) for t, v in (previous or {}).items()}

    def _age_days(stamp: str) -> int:
        try:
            return (datetime.strptime(today, "%Y-%m-%d")
                    - datetime.strptime(stamp, "%Y-%m-%d")).days
        except Exception:
            return 9999

    fetched, seen = 0, set()
    for t in tickers:
        tu = (t or "").upper()
        if not tu or tu in seen:
            continue
        seen.add(tu)
        cur = out.get(tu)
        if cur and cur.get("fetched") and _age_days(cur["fetched"]) < COMPANY_INFO_MAX_AGE_DAYS:
            continue
        if fetched >= COMPANY_INFO_MAX_FETCH_PER_RUN:
            continue
        try:
            info = yf.Ticker(t).info or {}
        except Exception as e:
            print(f"    company_info: {t} misslyckades ({e})")
            continue
        if not info:
            continue
        out[tu] = {
            "description": _trim_sentences(info.get("longBusinessSummary") or "", 3),
            "marketCap": info.get("marketCap"),
            "trailingPE": info.get("trailingPE"),
            "priceToSales": info.get("priceToSalesTrailing12Months"),
            "netMargin": info.get("profitMargins"),
            "revenueGrowth": info.get("revenueGrowth"),
            # trailingAnnualDividendYield is a clean fraction (0.0013); the raw
            # dividendYield field flip-flops between fraction and percent by
            # yfinance version, so avoid it.
            "dividendYield": info.get("trailingAnnualDividendYield"),
            "currency": info.get("currency"),
            "fetched": today,
        }
        fetched += 1
        print(f"    company_info: hämtade {t}")

    keep = {(t or "").upper() for t in tickers}
    out = {t: v for t, v in out.items() if t in keep}
    print(f"  company_info: {fetched} nya/uppdaterade, {len(out)} bolag totalt")
    return out


def _read_private_market_data() -> dict:
    """Läs förra körningens market_data.json från det PRIVATA repot — källan efter
    att den publika gisten togs bort. Används för carry-forward (portföljvärde,
    nyheter, kalender). Returnerar {} vid fel eller lokalt läge (ingen token)."""
    token = os.environ.get("PRIVATE_WRITE_TOKEN") or os.environ.get("DATA_TOKEN")
    if not token:
        return {}
    try:
        r = requests.get(
            "https://api.github.com/repos/linuspaulsson2/market-widget/contents/market_data.json?ref=main",
            headers={"Authorization": f"Bearer {token}", "Accept": "application/vnd.github.raw"},
            timeout=20,
        )
        return r.json() if r.status_code == 200 else {}
    except Exception:
        return {}


# Fält som avslöjar BELOPP — får aldrig hamna i den publika feeden.
AMOUNT_KEYS = ("shares", "invested", "value_sek", "pct_of_portfolio", "change_sek")
PUBLIC_GIST_ID = "b2b5723bb0a7396253041591548e413b"


def split_public_private(data: dict) -> tuple:
    """Dela feeden i (public_data, private_data). Publik = allt UTOM belopp:
    hela `portfolio`-objektet tas bort och AMOUNT_KEYS strippas ur varje
    all_holdings-post. Privat = { portfolio, amounts:{ticker:{belopp}} }."""
    import copy
    public = copy.deepcopy(data)
    public.pop("portfolio", None)
    for h in public.get("all_holdings", []) or []:
        for k in AMOUNT_KEYS:
            h.pop(k, None)

    amounts = {}
    for h in data.get("all_holdings", []) or []:
        t = h.get("ticker")
        if not t:
            continue
        amounts[t] = {k: h[k] for k in AMOUNT_KEYS if k in h}
    private = {
        "updated_at": data.get("updated_at", ""),
        "portfolio": data.get("portfolio", {}),
        "amounts": amounts,
    }
    return public, private


def _gh_put_private(path: str, content_str: str, message: str):
    """PUT en fil till det PRIVATA repot (contents-API). sys.exit(1) vid fel."""
    import base64
    token = os.environ.get("PRIVATE_WRITE_TOKEN") or os.environ.get("DATA_TOKEN")
    if not token:
        print(f"  (privat skrivning av {path} hoppas över — ingen token)")
        return
    api = f"https://api.github.com/repos/linuspaulsson2/market-widget/contents/{path}"
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
    try:
        sha = None
        g = requests.get(f"{api}?ref=main", headers=headers, timeout=30)
        if g.status_code == 200:
            sha = g.json().get("sha")
        body = {"message": message, "branch": "main",
                "content": base64.b64encode(content_str.encode("utf-8")).decode("ascii")}
        if sha:
            body["sha"] = sha
        p = requests.put(api, headers=headers, json=body, timeout=60)
        if p.status_code in (200, 201):
            print(f"  Privat {path} skriven")
        else:
            print(f"\nFEL: privat skrivning av {path}: HTTP {p.status_code} — {p.text[:200]}")
            sys.exit(1)
    except SystemExit:
        raise
    except Exception as e:
        print(f"\nFEL: privat skrivning av {path}: {e}")
        sys.exit(1)


def _read_arenas_private() -> str:
    """Läs arenas.json (JSON-sträng) från det privata repot för publik spegling."""
    token = os.environ.get("PRIVATE_WRITE_TOKEN") or os.environ.get("DATA_TOKEN")
    if not token:
        return ""
    try:
        r = requests.get(
            "https://api.github.com/repos/linuspaulsson2/market-widget/contents/arenas/arenas.json?ref=main",
            headers={"Authorization": f"Bearer {token}", "Accept": "application/vnd.github.raw"},
            timeout=30)
        return r.text if r.status_code == 200 else ""
    except Exception:
        return ""


def publish_public_gist(public_data: dict, arenas_json: str):
    """Uppdatera den PUBLIKA gisten (market_data.json utan belopp + arenas.json).
    Best-effort: en saknad gist-token loggar men fäller inte körningen."""
    token = os.environ.get("GIST_TOKEN") or os.environ.get("PUBLIC_GIST_TOKEN")
    if not token:
        print("  (publik gist hoppas över — ingen GIST_TOKEN med gist-scope)")
        return
    files = {"market_data.json": {"content": _safe_json(public_data, indent=2)}}
    if arenas_json:
        files["arenas.json"] = {"content": arenas_json}
    try:
        r = requests.patch(f"https://api.github.com/gists/{PUBLIC_GIST_ID}",
                           headers={"Authorization": f"token {token}",
                                    "Accept": "application/vnd.github+json"},
                           json={"files": files}, timeout=60)
        if r.status_code == 200:
            print("  Publik gist uppdaterad (utan belopp)")
        else:
            print(f"  (publik gist misslyckades: HTTP {r.status_code} — {r.text[:160]})")
    except Exception as e:
        print(f"  (publik gist fel: {e})")


def write_private_data(data: dict):
    """Skriv market_data.json till det PRIVATA repot (autentiserat) så portföljen
    aldrig ligger publikt läsbar. Best-effort: en saknad eller otillräcklig token
    loggar en varning men FÄLLER INTE körningen — så dubbelskrivningen under
    övergången aldrig kan bryta pipelinen. Kräver en token med Contents: Write på
    linuspaulsson2/market-widget (t.ex. DATA_TOKEN uppgraderad, eller en separat
    PRIVATE_WRITE_TOKEN)."""
    import base64
    token = os.environ.get("PRIVATE_WRITE_TOKEN") or os.environ.get("DATA_TOKEN")
    if not token:
        # Lokalt läge (ingen token): spara till fil istället för att fela.
        output_path = SCRIPT_DIR / "widget_data.json"
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(_safe_json(data, indent=2))
        print(f"\nIngen token — sparade data lokalt till: {output_path}")
        return
    api = "https://api.github.com/repos/linuspaulsson2/market-widget/contents/market_data.json"
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
    try:
        sha = None
        g = requests.get(f"{api}?ref=main", headers=headers, timeout=30)
        if g.status_code == 200:
            sha = g.json().get("sha")
        elif g.status_code != 404:
            print(f"  (privat skrivning: kunde inte läsa SHA, HTTP {g.status_code})")
        body = {
            "message": f"Update market_data.json ({data.get('updated_at', '')})",
            "content": base64.b64encode(_safe_json(data, indent=2).encode("utf-8")).decode("ascii"),
            "branch": "main",
        }
        if sha:
            body["sha"] = sha
        p = requests.put(api, headers=headers, json=body, timeout=60)
        if p.status_code in (200, 201):
            print("  Privat market_data.json skriven till linuspaulsson2/market-widget")
        else:
            print(f"\nFEL: privat skrivning misslyckades: HTTP {p.status_code} — {p.text[:200]}")
            sys.exit(1)
    except SystemExit:
        raise
    except Exception as e:
        print(f"\nFEL: privat skrivning fel: {e}")
        sys.exit(1)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print(f"Market Widget Update - {datetime.now(TZ_STOCKHOLM).strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # Kolla marknadsstatus
    market = is_market_open()
    print(f"\nMarknadsstatus: Stockholm={'Öppen' if market['stockholm'] else 'Stängd'}, "
          f"USA={'Öppen' if market['us'] else 'Stängd'}")

    # Hämta data
    print("\nHämtar aktiedata (Top Holdings)...")
    stocks = fetch_stock_data()

    print("\nHämtar aktiedata (Focus Stocks)...")
    focus_data = fetch_stock_data(FOCUS_STOCKS) if FOCUS_STOCKS else []

    print("\nHämtar USD/SEK...")
    fx = fetch_fx_data()
    print(f"  USD/SEK = {fx['usd_sek']} ({fx['usd_sek_change_pct']:+.2f}%)")

    print("\nHämtar index (S&P 500, OMXS30)...")
    indices = fetch_indices()

    print("\nHämtar Bitcoin & Guld...")
    commodities = fetch_commodities()

    print("\nHämtar Fear & Greed Index...")
    fear_greed = fetch_fear_greed()

    print("\nHämtar portföljvärde från Excel...")
    portfolio = fetch_portfolio_value()
    if portfolio:
        prev = fetch_previous_portfolio()
        if prev.get("prev_close_sek"):
            portfolio["prev_close_sek"] = prev["prev_close_sek"]
            change = portfolio["value_sek"] - prev["prev_close_sek"]
            change_pct = (change / prev["prev_close_sek"] * 100) if prev["prev_close_sek"] else 0
            portfolio["change_sek"] = round(change)
            portfolio["change_pct"] = round(change_pct, 2)
            print(f"  Förändring: {change:+,.0f} kr ({change_pct:+.2f}%)")
        else:
            portfolio["prev_close_sek"] = portfolio["value_sek"]
            portfolio["change_sek"] = 0
            portfolio["change_pct"] = 0.0
            print(f"  Första körningen — ingen föregående data")

        # Spara portfolio history för sparkline (max 30 datapunkter)
        old_history = prev.get("history", []) if prev else []
        today_str = datetime.now(TZ_STOCKHOLM).strftime("%Y-%m-%d")
        # Uppdatera dagens värde eller lägg till nytt
        updated = False
        for h in old_history:
            if h.get("date") == today_str:
                h["value"] = portfolio["value_sek"]
                updated = True
                break
        if not updated:
            old_history.append({"date": today_str, "value": portfolio["value_sek"]})
        # Behåll max 30 dagar
        portfolio["history"] = old_history[-30:]

    # Bygg alerts
    alerts = build_alerts(stocks)
    # Prepend split alerts: Yahoo reported a split not yet in splits.json, so this
    # holding's share count/value/return is wrong until a splits.json entry is added.
    for sp in _UNHANDLED_SPLITS:
        ratio = sp["ratio"]
        ratio_str = f"{int(ratio)}:1" if float(ratio).is_integer() else f"{ratio:g}:1"
        alerts.insert(0, {
            "type": "split",
            "severity": "high",
            "message": f"{sp['ticker']} gjorde en {ratio_str}-split {sp['date']} — lägg till i splits.json för rätt antal/värde",
            "ticker": sp["ticker"],
        })
    if alerts:
        print(f"\nAlerts ({len(alerts)} st):")
        for a in alerts:
            print(f"  [{a['severity']}] {a['message']}")

    # --- Nyhets-gating: nyheter/kalender är långsamma (Claude + RSS + .info).
    # NEWS_ENABLED=False stänger av dem HELT (tomma sektioner, ingen hämtning) —
    # tillfälligt av juli 2026 för att spara Actions-minuter. Sätt True för att
    # slå på igen (då hämtas de som mest var NEWS_REFRESH_HOURS:e timme).
    NEWS_ENABLED = False
    NEWS_REFRESH_HOURS = 3
    _prev_gist = _read_private_market_data()
    refresh_news = True
    _prev_news_stamp = _prev_gist.get("news_updated_at", "")
    try:
        _now_dt = datetime.strptime(market["timestamp_swe"], "%Y-%m-%d %H:%M")
        if _prev_news_stamp:
            _age_h = (_now_dt - datetime.strptime(_prev_news_stamp, "%Y-%m-%d %H:%M")).total_seconds() / 3600.0
            refresh_news = _age_h >= NEWS_REFRESH_HOURS
    except Exception:
        refresh_news = True
    news_updated_at = market["timestamp_swe"] if refresh_news else _prev_news_stamp
    print("\nNyheter: " + ("AVSTÄNGDA (sparar minuter)" if not NEWS_ENABLED
          else ("HÄMTAR färska" if refresh_news else "återanvänder cache från " + _prev_news_stamp)))

    if NEWS_ENABLED and refresh_news:
        # Samla bolagsnyheter (bara genuint viktiga)
        top_news = collect_top_news(stocks)
        print(f"\nBolagsnyheter: {len(top_news)} st")
        for n in top_news:
            print(f"  [{n['ticker']}] {n['title'][:70]}")

        # Hämta övergripande marknadsnyheter (S&P 500, OMXS30) — bara viktiga
        print("\nHämtar marknadsnyheter...")
        market_news = fetch_market_news()
        # Ta bort duplikater som redan finns i bolagsnyheter
        company_titles = {n["title"] for n in top_news}
        market_news = [n for n in market_news if n["title"] not in company_titles]
        print(f"  Marknadsnyheter: {len(market_news)} st")
        for n in market_news:
            print(f"  [{n.get('source', 'MKT')}] {n['title'][:70]}")

        # Injicera schemalagda händelser (FOMC, Triple Witching, etc.)
        # Dessa skickas INTE genom Claude-summarization — redan på svenska
        scheduled_events = _get_scheduled_events_for_today()
        if scheduled_events:
            print(f"  Schemalagda händelser: {len(scheduled_events)} st")
            for ev in scheduled_events:
                print(f"  📅 {ev['title']}")

        # AI-sammanfattning till korta svenska rubriker
        if top_news or market_news:
            print("\nSammanfattar nyheter med Claude...")
            top_news, market_news = summarize_news(top_news, market_news)
            for n in top_news:
                sv = n.get("title_sv", "")
                if sv:
                    print(f"  → {sv}")
            for n in market_news:
                sv = n.get("title_sv", "")
                if sv:
                    print(f"  → {sv}")

        # Lägg till schemalagda händelser EFTER summarization (de är redan på svenska)
        if scheduled_events:
            # Scheduled events först, sedan API-nyheter
            market_news = scheduled_events + market_news

        # Kolla kommande rapporter (90 dagar framåt, alla aktier)
        print("\nKollar kommande rapporter...")
        all_holdings = HOLDINGS + [h for h in FOCUS_STOCKS if h["ticker"] not in {x["ticker"] for x in HOLDINGS}]
        upcoming_earnings = fetch_upcoming_earnings(all_holdings, days_ahead=90)
        if upcoming_earnings:
            print(f"  Rapporterar denna vecka:")
            for e in upcoming_earnings:
                print(f"    {e['name']} - {e['date_display']}")
        else:
            print("  Inga rapporter denna vecka")

        # Injicera rapport-påminnelser som nyheter (inom 7 dagar)
        earnings_soon = [e for e in upcoming_earnings if e["days_until"] <= 7]
        for e in earnings_soon:
            days = e["days_until"]
            if days == 0:
                label = f"IDAG: {e['name']} rapporterar ({e['date_display']})"
            elif days == 1:
                label = f"IMORGON: {e['name']} rapporterar ({e['date_display']})"
            else:
                label = f"OM {days} DAGAR: {e['name']} rapporterar ({e['date_display']})"
            ticker = e["ticker"]
            is_swedish = ticker.endswith(".ST")
            top_news.append({
                "title": label,
                "title_sv": label,
                "publisher": "Kalender",
                "ticker": ticker,
                "stock_name": e["name"],
                "signal_type": "RAPPORT",
                "impact": "HIGH" if days <= 1 else "MED",
            })
            print(f"  📅 Rapport-påminnelse: {label}")

        # Hämta kommande utdelningar
        print("\nKollar kommande utdelningar...")
        upcoming_dividends = fetch_upcoming_dividends(all_holdings, days_ahead=90)
        if upcoming_dividends:
            print(f"  Utdelningar kommande 90 dagar:")
            for d in upcoming_dividends:
                amt = f"{d['amount']} {d['currency']}" if d.get('amount') else "?"
                print(f"    {d['name']} — ex {d['ex_date_display']} — {amt}")
        else:
            print("  Inga kommande utdelningar")

    elif NEWS_ENABLED:
        top_news = _prev_gist.get("news", []) or []
        market_news = _prev_gist.get("market_news", []) or []
        upcoming_earnings = _prev_gist.get("upcoming_earnings", []) or []
        upcoming_dividends = _prev_gist.get("upcoming_dividends", []) or []
        print(f"  Återanvänder {len(top_news)} nyheter, {len(market_news)} marknadsnyheter, {len(upcoming_earnings)} rapporter, {len(upcoming_dividends)} utdelningar")
    else:
        # Nyheter helt avstängda — tomma sektioner, ingen långsam hämtning.
        top_news, market_news, upcoming_earnings, upcoming_dividends = [], [], [], []
        news_updated_at = ""

    # Hämta alla holdings (hela portföljen)
    print("\nHämtar alla holdings (batch)...")
    all_holdings_data = fetch_all_holdings()

    # --- Kalender (rapporter + utdelningar) — OBEROENDE av NEWS_ENABLED. ---
    # Rapport-/utdelningsdatum ändras sällan → räkna om var CALENDAR_REFRESH_HOURS:e
    # timme, återanvänd annars förra gistens kalender så den ÖVERLEVER varje cron
    # (bugg innan: else-grenen ovan nollade kalendern varje körning). "Lita aldrig
    # på tomt": om cachen är tom räknar vi om ändå, så en enstaka stale gist-läsning
    # inte permanent-nollar kalendern. Full portfölj — inte bara config/focus.
    CALENDAR_REFRESH_HOURS = 12
    _prev_cal_e = _prev_gist.get("upcoming_earnings", []) or []
    _prev_cal_d = _prev_gist.get("upcoming_dividends", []) or []
    _prev_cal_stamp = _prev_gist.get("calendar_updated_at", "")
    _cal_stale = True
    try:
        if _prev_cal_stamp:
            _cage = (_now_dt - datetime.strptime(_prev_cal_stamp, "%Y-%m-%d %H:%M")).total_seconds() / 3600.0
            _cal_stale = _cage >= CALENDAR_REFRESH_HOURS
    except Exception:
        _cal_stale = True

    if _cal_stale or not _prev_cal_e:
        cal_src = [{"ticker": h["ticker"], "name": h.get("name", h["ticker"])}
                   for h in all_holdings_data
                   if h.get("ticker")
                   and not h["ticker"].upper().endswith("-USD")   # krypto
                   and not h["ticker"].upper().startswith("0P")]  # fonder
        print(f"\nKalender: räknar om för {len(cal_src)} innehav…")
        upcoming_earnings = fetch_upcoming_earnings(cal_src, days_ahead=90)
        upcoming_dividends = fetch_upcoming_dividends(cal_src, days_ahead=90)
        calendar_updated_at = market["timestamp_swe"]
        print(f"  {len(upcoming_earnings)} rapporter, {len(upcoming_dividends)} utdelningar")
    else:
        upcoming_earnings, upcoming_dividends = _prev_cal_e, _prev_cal_d
        calendar_updated_at = _prev_cal_stamp
        print(f"\nKalender: återanvänder {len(upcoming_earnings)} rapporter, "
              f"{len(upcoming_dividends)} utdelningar (färsk nog)")

    # Hämta watchlist-aktier (bevakning, ej köpt än). Exkludera tickers du redan
    # äger så att en bevakad aktie som blivit köpt bara visas under Innehav.
    print("\nHämtar watchlist (batch)...")
    held_now = {h["ticker"].upper() for h in all_holdings_data if h.get("shares", 0) > 0}
    watchlist_data = fetch_watchlist(held_now)

    # Portfolio value from the live holdings sum (the old Dropbox/Excel source is
    # gone). Ensures the gist's portfolio object is COMPLETE — the app requires
    # value_sek/change_sek/change_pct and an empty {} fails to decode.
    if all_holdings_data:
        portfolio["value_sek"] = round(sum(h.get("value_sek", 0) or 0 for h in all_holdings_data))

    # Räkna om portföljens dagliga SEK-förändring från live-holdings
    # (change_pct per holding är live)
    if portfolio and all_holdings_data:
        live_change_sek = sum(
            h.get("value_sek", 0) * h.get("change_pct", 0) / 100
            for h in all_holdings_data
        )
        total_val = portfolio.get("value_sek", 0)
        live_change_pct = (live_change_sek / (total_val - live_change_sek) * 100) if total_val else 0
        portfolio["change_sek"] = round(live_change_sek)
        portfolio["change_pct"] = round(live_change_pct, 2)
        print(f"  Daglig förändring (live): {live_change_sek:+,.0f} kr ({live_change_pct:+.2f}%)")

    # Guarantee a complete portfolio object even if holdings couldn't be fetched
    # (the app's Portfolio decoder requires all three fields; {} crashes it).
    portfolio.setdefault("value_sek", 0)
    portfolio.setdefault("change_sek", 0)
    portfolio.setdefault("change_pct", 0.0)

    # Sortera aktier efter störst rörelse (mest intressant först)
    stocks_sorted = sorted(stocks, key=lambda s: abs(s["change_pct"]), reverse=True)

    # Veckosammanfattning (bara helger)
    weekly_summary = build_weekly_summary(stocks, portfolio, indices)
    if weekly_summary:
        print(f"\nVeckosammanfattning: {weekly_summary.get('text', '')[:80]}...")

    # Bolagsinfo (beskrivning + 6 nyckeltal) för detaljvyn — cachad i gisten.
    # Får aldrig fälla hela körningen: vid fel behåller vi förra gistens cache.
    print("\nBolagsinfo (beskrivning + nyckeltal)...")
    _prev_company_info = _prev_gist.get("company_info", {}) or {}
    try:
        company_info = build_company_info(
            [h["ticker"] for h in all_holdings_data] + [w["ticker"] for w in watchlist_data],
            _prev_company_info,
            market["timestamp_swe"][:10],
        )
    except Exception as _e:
        print(f"  (company_info misslyckades, behåller cache: {_e})")
        company_info = _prev_company_info

    # Bygg slutlig JSON
    data = {
        "updated_at": market["timestamp_swe"],
        "news_updated_at": news_updated_at,
        "market_status": {
            "stockholm": "open" if market["stockholm"] else "closed",
            "us": "open" if market["us"] else "closed",
        },
        "fx": fx,
        "indices": indices,
        "commodities": commodities,
        "fear_greed": fear_greed,
        "portfolio": portfolio,
        "stocks": stocks_sorted,
        "focus_stocks": focus_data,
        "alerts": alerts,
        "news": top_news,
        "market_news": market_news,
        "upcoming_earnings": upcoming_earnings,
        "upcoming_dividends": upcoming_dividends,
        "calendar_updated_at": calendar_updated_at,
        "weekly_summary": weekly_summary,
        "all_holdings": all_holdings_data,
        "watchlist": watchlist_data,
        "company_info": company_info,
    }

    # Publicera i tre delar:
    #  - privat FULL market_data.json → carry-forward + installerad app under övergången
    #  - privat portfolio.json        → BELOPPEN (nya appen läser, aldrig publikt)
    #  - publik gist                  → allt UTOM belopp + arenas.json (spegel)
    public_data, private_data = split_public_private(data)
    write_private_data(data)
    _gh_put_private("portfolio.json", _safe_json(private_data, indent=2),
                    f"Update portfolio.json ({data.get('updated_at', '')})")
    publish_public_gist(public_data, _read_arenas_private())
    print("\nKlart!")


if __name__ == "__main__":
    main()
